# UI/dialogs.py
import datetime
import re
import webbrowser
import tkinter as tk
import tkinter.font as tkfont
from tkinter import filedialog, messagebox, simpledialog, ttk

from UI.date_picker import DatePicker

from utils import make_preview
from db import (
    conn,
    update_statusuri_din_rezervari,
    create_user,
    get_location_by_id,
    add_client_contact,
    get_client_contacts,
    update_client_contact,
    delete_client_contact,
    table_has_column,
)


def _safe_filename(name: str) -> str:
    """Return *name* sanitized for filesystem usage."""
    return re.sub(r"[\\/*?:\"<>|]", "_", name)


def choose_report_year(parent=None):
    """Return the year selected by the user for the sales report or ``None``."""
    current_year = datetime.date.today().year
    cur = conn.cursor()
    try:
        cur.execute("SELECT MIN(data_start), MAX(data_end) FROM rezervari")
        row = cur.fetchone()
    except Exception:
        row = None
    min_year = max_year = current_year
    if row:
        dmin, dmax = row
        try:
            if dmin:
                if isinstance(dmin, str):
                    min_year = datetime.date.fromisoformat(dmin).year
                else:
                    min_year = dmin.year
        except Exception:
            pass
        try:
            if dmax:
                if isinstance(dmax, str):
                    max_year = datetime.date.fromisoformat(dmax).year
                else:
                    max_year = dmax.year
        except Exception:
            pass
    years = sorted({*range(min_year, max_year + 1), current_year})

    win = tk.Toplevel(parent)
    win.title("An raport vânzări")
    ttk.Label(win, text="An:").grid(row=0, column=0, padx=5, pady=5)
    year_var = tk.StringVar(value=str(current_year))
    cb = ttk.Combobox(win, textvariable=year_var, values=[str(y) for y in years], state="readonly", width=8)
    cb.grid(row=0, column=1, padx=5, pady=5)

    result = {"year": None}

    def ok():
        try:
            result["year"] = int(year_var.get())
        except Exception:
            result["year"] = current_year
        win.destroy()

    def cancel():
        win.destroy()

    ttk.Button(win, text="OK", command=ok).grid(row=1, column=0, padx=5, pady=5)
    ttk.Button(win, text="Renunță", command=cancel).grid(row=1, column=1, padx=5, pady=5)

    win.grab_set()
    win.wait_window()
    return result["year"]


def choose_report_type(role: str, parent=None) -> str | None:
    """Return the report type selected by the user."""
    options = ["Vânzări", "Decorări"]
    if role in ("admin", "manager"):
        options.append("Vânzători")

    win = tk.Toplevel(parent)
    win.title("Tip raport")
    ttk.Label(win, text="Raport:").grid(row=0, column=0, padx=5, pady=5)
    var = tk.StringVar(value=options[0])
    cb = ttk.Combobox(win, values=options, textvariable=var, state="readonly")
    cb.grid(row=0, column=1, padx=5, pady=5)

    result = {"value": None}

    def ok():
        result["value"] = var.get()
        win.destroy()

    def cancel():
        win.destroy()

    ttk.Button(win, text="OK", command=ok).grid(row=1, column=0, padx=5, pady=5)
    ttk.Button(win, text="Renunță", command=cancel).grid(row=1, column=1, padx=5, pady=5)
    win.grab_set()
    win.wait_window()
    return result["value"]


def open_detail_window(tree, event):
    """Display extended information about the selected location."""
    rowid = tree.identify_row(event.y)
    if not rowid:
        return
    loc_id = int(rowid)  # folosim iid-ul (id-ul real din DB), nu valoarea din coloane
    # extrage detaliile locației și le afișează într-o fereastră dedicată

    data = get_location_by_id(loc_id)
    if not data:
        messagebox.showerror("Eroare", "Datele locației nu au fost găsite.")
        return

    city = data.get("city")
    county = data.get("county")
    address = data.get("address")
    type_ = data.get("type")
    gps = data.get("gps")
    code = data.get("code")
    size_ = data.get("size")
    photo_link = data.get("photo_link")
    sqm = data.get("sqm")
    illumination = data.get("illumination")
    ratecard = data.get("ratecard")
    decoration_cost = data.get("decoration_cost")
    pret_vanzare = data.get("pret_vanzare")
    pret_flotant = data.get("pret_flotant")
    observatii = data.get("observatii")
    status = data.get("status")
    client = data.get("client")
    ds = data.get("data_start")
    de = data.get("data_end")
    cur = conn.cursor()

    # fereastra
    win = tk.Toplevel(tree.master)
    win.title(f"Detalii locație #{loc_id}")

    frm = ttk.Frame(win, padding=10)
    frm.grid(sticky="nsew")
    win.columnconfigure(0, weight=1)
    win.rowconfigure(0, weight=1)

    # 1) Preview imagine
    img = make_preview(code)
    if img:
        lbl_img = ttk.Label(frm, image=img)
        lbl_img.image = img
    else:
        lbl_img = ttk.Label(frm, text="Fără preview")
    lbl_img.grid(row=0, column=0, columnspan=2, pady=(0, 15))

    # helper pentru rânduri
    def add_field(r, label, widget):
        ttk.Label(frm, text=label + ":", font=("Segoe UI", 9, "bold")).grid(
            row=r, column=0, sticky="e", padx=5, pady=2
        )
        widget.grid(row=r, column=1, sticky="w", padx=5, pady=2)

    r = 1
    add_field(r, "City", ttk.Label(frm, text=city))
    r += 1
    add_field(r, "County", ttk.Label(frm, text=county))
    r += 1
    add_field(r, "Address", ttk.Label(frm, text=address))
    r += 1
    add_field(r, "Type", ttk.Label(frm, text=type_))
    r += 1

    # GPS ca hyperlink către Google Maps, cu text "Google Maps"
    if gps:
        url_maps = f"https://www.google.com/maps/search/?api=1&query={gps}"
        lbl_gps = ttk.Label(frm, text="Google Maps", cursor="hand2", foreground="blue")
        lbl_gps.bind("<Button-1>", lambda e: webbrowser.open(url_maps))
    else:
        lbl_gps = ttk.Label(frm, text="-")
    add_field(r, "GPS", lbl_gps)
    r += 1

    add_field(r, "Code", ttk.Label(frm, text=code))
    r += 1
    add_field(r, "Size", ttk.Label(frm, text=size_))
    r += 1

    # Photo Link ca hyperlink
    if photo_link:
        href = photo_link.strip()
        if not href.lower().startswith(("http://", "https://")):
            href = "https://" + href
        lbl_photo = ttk.Label(frm, text="Vezi poza", cursor="hand2", foreground="blue")
        lbl_photo.bind("<Button-1>", lambda e: webbrowser.open(href))
    else:
        lbl_photo = ttk.Label(frm, text="-")
    add_field(r, "Photo Link", lbl_photo)
    r += 1

    add_field(r, "SQM", ttk.Label(frm, text=str(sqm)))
    r += 1
    add_field(r, "Illumination", ttk.Label(frm, text=illumination))
    r += 1
    add_field(r, "RateCard", ttk.Label(frm, text=str(ratecard)))
    r += 1
    add_field(r, "Preț de vânzare", ttk.Label(frm, text=str(pret_vanzare)))
    r += 1
    add_field(r, "Preț Flotant", ttk.Label(frm, text=str(pret_flotant)))
    r += 1
    add_field(r, "Preț de decorare", ttk.Label(frm, text=str(decoration_cost)))
    r += 1

    add_field(r, "Observații", ttk.Label(frm, text=observatii or "-"))
    r += 1
    add_field(r, "Status", ttk.Label(frm, text=status))
    r += 1

    # Client și perioadă doar dacă există
    if client:
        add_field(r, "Client", ttk.Label(frm, text=client))
        r += 1
        period = f"{ds} → {de}" if ds and de else "-"
        add_field(r, "Perioadă", ttk.Label(frm, text=period))
        r += 1
        if ds and de:
            row_info = cur.execute(
                "SELECT created_by, suma, campaign FROM rezervari WHERE loc_id=? AND data_start=? AND data_end=? ORDER BY data_start DESC LIMIT 1",
                (loc_id, ds, de),
            ).fetchone()
        else:
            row_info = None
        if row_info:
            created_by, suma_val, campaign_val = row_info
            if suma_val is not None:
                add_field(r, "Sumă închiriere", ttk.Label(frm, text=str(suma_val)))
                r += 1
            if created_by:
                add_field(r, "Închiriat de", ttk.Label(frm, text=created_by))
                r += 1
            if campaign_val:
                add_field(r, "Campanie", ttk.Label(frm, text=campaign_val))
                r += 1

    # face fereastra redimensionabilă
    for i in range(r):
        win.rowconfigure(i, weight=0)
    win.columnconfigure(1, weight=1)


def open_add_window(root, refresh_cb):
    win = tk.Toplevel(root)
    win.title("Adaugă locație")
    labels = [
        "City",
        "County",
        "Address",
        "Type",
        "GPS",
        "Code",
        "Size",
        "Photo Link",
        "SQM",
        "Illumination",
        "RateCard",
        "Preț Vânzare",
        "Preț Flotant",
        "Decoration cost",
        "Observații",
        "Grup",
        "Față",
    ]
    entries = {}
    for i, lbl in enumerate(labels):
        ttk.Label(win, text=lbl + ":").grid(row=i, column=0, sticky="e", padx=5, pady=2)
        if lbl == "Față":
            e = ttk.Combobox(win, values=["Fața A", "Fața B"], state="readonly")
            e.current(0)
        else:
            e = ttk.Entry(win, width=40)
        e.grid(row=i, column=1, padx=5, pady=2)
        entries[lbl] = e

    var_mobile = tk.BooleanVar()
    ttk.Checkbutton(win, text="Este locație mobilă", variable=var_mobile).grid(
        row=len(labels), column=0, columnspan=2, pady=(5, 0)
    )

    def save():
        vals = {k: e.get().strip() for k, e in entries.items()}
        if not vals["City"]:
            messagebox.showwarning("Lipsește date", "Completează City.")
            return
        if var_mobile.get():
            vals.setdefault("Grup", "Prisme mobile")
            vals.setdefault("Type", "Prismă mobilă")
            vals.setdefault("Size", "3.2x2.4")
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO locatii (
                city, county, address, type, gps, code, size,
                photo_link, sqm, illumination, ratecard,
                pret_vanzare, pret_flotant, decoration_cost,
                observatii, grup, face, is_mobile, parent_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            [
                vals["City"],
                vals["County"],
                vals["Address"],
                vals["Type"],
                vals["GPS"],
                vals["Code"],
                vals["Size"],
                vals["Photo Link"],
                vals["SQM"],
                vals["Illumination"],
                vals["RateCard"],
                vals["Preț Vânzare"] or None,
                vals["Preț Flotant"] or None,
                vals["Decoration cost"] or None,
                vals["Observații"],
                vals["Grup"],
                vals["Față"],
                1 if var_mobile.get() else 0,
                None,
            ],
        )
        conn.commit()
        refresh_cb()
        win.destroy()

    ttk.Button(win, text="Salvează", command=save).grid(
        row=len(labels) + 1, column=0, columnspan=2, pady=10
    )


def open_edit_window(root, loc_id, load_cb, refresh_groups_cb):
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            city, county, address, type, gps, code, size,
            photo_link, sqm, illumination, ratecard,
            pret_vanzare, pret_flotant, decoration_cost,
            observatii, grup, face
        FROM locatii WHERE id=?
    """,
        (loc_id,),
    )
    row = cur.fetchone()
    if not row:
        return

    labels = [
        "City",
        "County",
        "Address",
        "Type",
        "GPS",
        "Code",
        "Size",
        "Photo Link",
        "SQM",
        "Illumination",
        "RateCard",
        "Preț Vânzare",
        "Preț Flotant",
        "Decoration cost",
        "Observații",
        "Grup",
        "Față",
    ]

    win = tk.Toplevel(root)
    win.title(f"Editează locație #{loc_id}")
    entries = {}

    # Asigură-te că row are aceeași lungime cu labels
    row = list(row)
    while len(row) < len(labels):
        row.append("")

    for i, lbl in enumerate(labels):
        ttk.Label(win, text=lbl + ":").grid(row=i, column=0, sticky="e", padx=5, pady=2)
        if lbl == "Față":
            e = ttk.Combobox(win, values=["Fața A", "Fața B"], state="readonly")
            e.set(row[i] if row[i] else "Fața A")
        else:
            e = ttk.Entry(win, width=40)
            e.insert(0, str(row[i]) if row[i] is not None else "")
        e.grid(row=i, column=1, padx=5, pady=2)
        entries[lbl] = e

    def save_edit():
        vals = {k: e.get().strip() for k, e in entries.items()}
        if not vals["City"]:
            messagebox.showwarning("Lipsește date", "Completează City.")
            return
        cur = conn.cursor()
        cur.execute(
            """
            UPDATE locatii SET
                city=?, county=?, address=?, type=?, gps=?, code=?, size=?,
                photo_link=?, sqm=?, illumination=?, ratecard=?,
                pret_vanzare=?, pret_flotant=?, decoration_cost=?,
                observatii=?, grup=?, face=?
            WHERE id=?
        """,
            [
                vals["City"],
                vals["County"],
                vals["Address"],
                vals["Type"],
                vals["GPS"],
                vals["Code"],
                vals["Size"],
                vals["Photo Link"],
                vals["SQM"],
                vals["Illumination"],
                vals["RateCard"],
                vals["Preț Vânzare"] or None,
                vals["Preț Flotant"] or None,
                vals["Decoration cost"] or None,
                vals["Observații"],
                vals["Grup"],
                vals["Față"],
                loc_id,
            ],
        )
        conn.commit()
        refresh_groups_cb()
        load_cb()
        win.destroy()

    ttk.Button(win, text="Salvează modificările", command=save_edit).grid(
        row=len(labels), column=0, columnspan=2, pady=10
    )


def cancel_reservation(root, loc_id, load_cb):
    if not messagebox.askyesno(
        "Confirmă anulare",
        "Sigur vrei să anulezi rezervarea/închirierea acestei locații?",
    ):
        return

    cur = conn.cursor()
    # Resetăm câmpurile în locatii
    cur.execute(
        """
        UPDATE locatii
           SET status     = 'Disponibil',
               client     = NULL,
               data_start = NULL,
               data_end   = NULL
         WHERE id = ?
    """,
        (loc_id,),
    )

    # Ștergem decorările legate de rezervările acestei locații
    rez_ids = [r[0] for r in cur.execute(
        "SELECT id FROM rezervari WHERE loc_id=?",
        (loc_id,),
    ).fetchall()]
    for rid in rez_ids:
        cur.execute("DELETE FROM decorari WHERE rez_id=?", (rid,))

    # Ștergem și intrările din tabelul rezervari
    cur.execute("DELETE FROM rezervari WHERE loc_id=?", (loc_id,))

    conn.commit()

    load_cb()


def open_reserve_window(root, loc_id, load_cb, user):
    """Rezervă locația pentru 5 zile folosind doar numele clientului."""
    loc_data = get_location_by_id(loc_id)
    if loc_data and loc_data.get("is_mobile") and not loc_data.get("parent_id"):
        messagebox.showwarning(
            "Refuzat",
            "Prismele mobile nu pot fi rezervate. Folosește închirierea.",
        )
        return

    win = tk.Toplevel(root)
    win.title(f"Rezervă locația #{loc_id}")

    ttk.Label(win, text="Client:").grid(row=0, column=0, padx=5, pady=5)
    entry_client = ttk.Entry(win, width=30)
    entry_client.grid(row=0, column=1, padx=5, pady=5)

    def save():
        name = entry_client.get().strip()
        if not name:
            messagebox.showwarning("Lipsește client", "Completează numele clientului.")
            return
        start = datetime.date.today()
        end = start + datetime.timedelta(days=4)
        cur = conn.cursor()
        created_on = datetime.date.today().isoformat()
        cur.execute(
            "INSERT INTO rezervari (loc_id, client, data_start, data_end, created_by, created_on) VALUES (?, ?, ?, ?, ?, ?)",
            (
                loc_id,
                name,
                start.isoformat(),
                end.isoformat(),
                user["username"],
                created_on,
            ),
        )
        conn.commit()
        update_statusuri_din_rezervari()
        load_cb()
        win.destroy()

    ttk.Button(win, text="Rezervă", command=save).grid(
        row=1, column=0, columnspan=2, pady=10
    )


def open_rent_window(root, loc_id, load_cb, user):
    """Dialog pentru adăugarea unei închirieri în tabelul ``rezervari``.

    Perioada aleasă trebuie să nu se suprapună peste o rezervare sau o
    închiriere existentă pentru aceeași locație. După salvare statusurile sunt
    recalculate prin ``update_statusuri_din_rezervari``.
    """

    win = tk.Toplevel(root)
    win.title(f"Închiriază locația #{loc_id}")

    loc_data = get_location_by_id(loc_id)
    is_mobile = loc_data.get("is_mobile") if loc_data else 0
    is_base_mobile = is_mobile and not loc_data.get("parent_id")

    ttk.Label(win, text="Client:").grid(row=0, column=0, sticky="e", padx=5, pady=5)

    def client_list():
        return [
            r[0]
            for r in conn.cursor()
            .execute("SELECT nume FROM clienti ORDER BY nume")
            .fetchall()
        ]

    cb_client = ttk.Combobox(win, values=client_list(), width=27)
    cb_client.grid(row=0, column=1, padx=5, pady=5)
    ttk.Button(
        win,
        text="+",
        command=lambda: (
            open_add_client_window(
                win, lambda: cb_client.configure(values=client_list())
            )
        ),
    ).grid(row=0, column=2, padx=2, pady=5)

    ttk.Label(win, text="Societate:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    def firma_list():
        return [r[0] for r in conn.cursor().execute("SELECT nume FROM firme ORDER BY nume").fetchall()]
    cb_firma = ttk.Combobox(win, values=firma_list(), width=27)
    cb_firma.grid(row=1, column=1, padx=5, pady=5)
    ttk.Button(
        win,
        text="+",
        command=lambda: (open_firme_window(win), cb_firma.configure(values=firma_list()))
    ).grid(row=1, column=2, padx=2, pady=5)
    labels = ["Data start", "Data end", "Sumă finală"]
    entries = {}
    for i, lbl in enumerate(labels, start=2):
        ttk.Label(win, text=lbl + ":").grid(row=i, column=0, sticky="e", padx=5, pady=5)
        if "Data" in lbl:
            e = DatePicker(win)
        else:
            e = ttk.Entry(win, width=30)
        e.grid(row=i, column=1, padx=5, pady=5)
        entries[lbl] = e

    row_next = len(labels) + 2
    ttk.Label(win, text="Preț decorare:").grid(row=row_next, column=0, sticky="e", padx=5, pady=5)
    entry_deco = ttk.Entry(win, width=30)
    deco_default = loc_data.get("decoration_cost") if loc_data else None
    if deco_default:
        entry_deco.insert(0, str(deco_default))
    entry_deco.grid(row=row_next, column=1, padx=5, pady=5)

    row_next += 1
    var_prod = tk.BooleanVar(value=False)
    chk_prod = ttk.Checkbutton(win, text="Add production cost", variable=var_prod, command=lambda: entry_prod.grid() if var_prod.get() else entry_prod.grid_remove())
    chk_prod.grid(row=row_next, column=0, sticky="e", padx=5, pady=5)
    entry_prod = ttk.Entry(win, width=30)
    sqm_val = loc_data.get("sqm") if loc_data else 0
    try:
        prod_default = round(float(sqm_val or 0) * 7, 2)
    except Exception:
        prod_default = 0
    entry_prod.insert(0, str(prod_default))
    entry_prod.grid(row=row_next, column=1, padx=5, pady=5)
    entry_prod.grid_remove()

    row_campaign = row_next + 1
    lbl_camp = ttk.Label(win, text="Campanie:")
    entry_camp = ttk.Entry(win, width=30)
    lbl_camp.grid(row=row_campaign, column=0, sticky="e", padx=5, pady=5)
    entry_camp.grid(row=row_campaign, column=1, padx=5, pady=5)
    lbl_camp.grid_remove()
    entry_camp.grid_remove()

    def toggle_campaign(_=None):
        name = cb_client.get().strip()
        r = conn.cursor().execute("SELECT tip FROM clienti WHERE nume=?", (name,)).fetchone()
        if r and r[0] == "agency":
            lbl_camp.grid()
            entry_camp.grid()
        else:
            lbl_camp.grid_remove()
            entry_camp.grid_remove()
            entry_camp.delete(0, tk.END)

    cb_client.bind("<<ComboboxSelected>>", toggle_campaign)
    toggle_campaign()

    row_extra = row_campaign + 1
    if is_base_mobile:
        ttk.Label(win, text="Adresă montaj:").grid(
            row=row_extra, column=0, sticky="e", padx=5, pady=5
        )
        entry_addr = ttk.Entry(win, width=30)
        entry_addr.grid(row=row_extra, column=1, padx=5, pady=5)
        row_extra += 1
        ttk.Label(win, text="GPS:").grid(
            row=row_extra, column=0, sticky="e", padx=5, pady=5
        )
        entry_gps = ttk.Entry(win, width=30)
        entry_gps.grid(row=row_extra, column=1, padx=5, pady=5)
        row_extra += 1
    else:
        entry_addr = None
        entry_gps = None

    def save_rent():
        client = cb_client.get().strip()
        if not client:
            messagebox.showwarning("Lipsește client", "Completează client.")
            return

        cur = conn.cursor()
        row = cur.execute(
            "SELECT id, tip FROM clienti WHERE nume=?", (client,)
        ).fetchone()
        if row:
            client_id, tip = row
        else:
            cur.execute("INSERT INTO clienti (nume) VALUES (?)", (client,))
            conn.commit()
            client_id = cur.lastrowid
            tip = "direct"

        campaign_val = entry_camp.get().strip() if lbl_camp.winfo_ismapped() else ""
        client_display = client
        if tip == "agency":
            if not campaign_val:
                messagebox.showwarning(
                    "Lipsește campania", "Completează denumirea campaniei."
                )
                return
            client_display = f"{client} - {campaign_val}"

        start = entries["Data start"].get_date()
        end = entries["Data end"].get_date()
        if start > end:
            messagebox.showwarning(
                "Interval incorect",
                "«Data start» trebuie înainte de «Data end».",
            )
            return

        fee_txt = entries["Sumă finală"].get().strip()
        try:
            fee_val = float(fee_txt)
        except ValueError:
            messagebox.showwarning("Sumă invalidă", "Introdu o sumă numerică.")
            return

        deco_txt = entry_deco.get().strip()
        try:
            deco_val = float(deco_txt)
        except ValueError:
            messagebox.showwarning("Decorare invalidă", "Introdu un cost numeric.")
            return

        if var_prod.get():
            prod_txt = entry_prod.get().strip()
            try:
                prod_val = float(prod_txt)
            except ValueError:
                messagebox.showwarning("Producție invalidă", "Introdu un cost numeric.")
                return
        else:
            prod_val = 0.0

        firma_name = cb_firma.get().strip()
        cur = conn.cursor()
        if firma_name:
            row_f = cur.execute("SELECT id FROM firme WHERE nume=?", (firma_name,)).fetchone()
            if row_f:
                firma_id = row_f[0]
            else:
                cur.execute("INSERT INTO firme (nume) VALUES (?)", (firma_name,))
                conn.commit()
                firma_id = cur.lastrowid
        else:
            firma_id = None

        # ``campaign_val`` already computed above

        cur = conn.cursor()

        # verificăm suprapuneri cu alte perioade
        rows = cur.execute(
            "SELECT suma FROM rezervari WHERE loc_id=? AND NOT (data_end < ? OR data_start > ?)",
            (loc_id, start.isoformat(), end.isoformat()),
        ).fetchall()
        for (suma,) in rows:
            # Orice închiriere existentă blochează intervalul. În cazul prismei
            # mobile, ignorăm suprapunerile cu înregistrări care au suma ``0``
            # (create pentru locația de bază).
            if suma is None or suma > 0:
                messagebox.showerror(
                    "Perioadă ocupată",
                    "Locația este deja rezervată sau închiriată în intervalul ales.",
                )
                return

        if is_base_mobile:
            cnt = cur.execute(
                """
                SELECT COUNT(*) FROM rezervari
                 WHERE loc_id IN (SELECT id FROM locatii WHERE parent_id=?)
                   AND NOT (data_end < ? OR data_start > ?)
                   AND suma IS NOT NULL
                """,
                (loc_id, start.isoformat(), end.isoformat()),
            ).fetchone()[0]
            if cnt >= 20:
                messagebox.showerror(
                    "Limită depășită",
                    "Nu poți închiria mai mult de 20 de prisme simultan.",
                )
                return

            addr_val = entry_addr.get().strip() if entry_addr else ""
            gps_val = entry_gps.get().strip() if entry_gps else ""
            if not addr_val or not gps_val:
                messagebox.showwarning(
                    "Lipsește adresa", "Completează adresa și GPS-ul."
                )
                return

            base = get_location_by_id(loc_id)
            cur.execute(
                """
                INSERT INTO locatii (
                    city, county, address, type, gps, code, size,
                    photo_link, sqm, illumination, ratecard,
                    pret_vanzare, pret_flotant, decoration_cost,
                    observatii, grup, face, is_mobile, parent_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                """,
                [
                    base.get("city"),
                    base.get("county"),
                    addr_val,
                    base.get("type"),
                    gps_val,
                    base.get("code"),
                    base.get("size"),
                    base.get("photo_link"),
                    base.get("sqm"),
                    base.get("illumination"),
                    base.get("ratecard"),
                    base.get("pret_vanzare"),
                    base.get("pret_flotant"),
                    base.get("decoration_cost"),
                    base.get("observatii"),
                    base.get("grup"),
                    base.get("face"),
                    loc_id,
                ],
            )
            new_loc_id = cur.lastrowid
            created_on = datetime.date.today().isoformat()
            cur.execute(
                "INSERT INTO rezervari (loc_id, client, client_id, firma_id, data_start, data_end, suma, created_by, created_on, campaign, decor_cost, prod_cost)"
                " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    new_loc_id,
                    client_display,
                    client_id,
                    firma_id,
                    start.isoformat(),
                    end.isoformat(),
                    fee_val,
                    user["username"],
                    created_on,
                    campaign_val or client_display,
                    deco_val,
                    prod_val,
                ),
            )
            rez_id = cur.lastrowid
            cur.execute(
                "INSERT INTO rezervari (loc_id, client, client_id, firma_id, data_start, data_end, suma, created_by, created_on, campaign, decor_cost, prod_cost)"
                " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    loc_id,
                    client_display,
                    client_id,
                    firma_id,
                    start.isoformat(),
                    end.isoformat(),
                    0.0,
                    user["username"],
                    created_on,
                    campaign_val or client_display,
                    0.0,
                    0.0,
                ),
            )
        else:
            # inserăm noua închiriere
            created_on = datetime.date.today().isoformat()
            cur.execute(
                "INSERT INTO rezervari (loc_id, client, client_id, firma_id, data_start, data_end, suma, created_by, created_on, campaign, decor_cost, prod_cost)"
                " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    loc_id,
                    client_display,
                    client_id,
                    firma_id,
                    start.isoformat(),
                    end.isoformat(),
                    fee_val,
                    user["username"],
                    created_on,
                    campaign_val or client_display,
                    deco_val,
                    prod_val,
                ),
            )
            rez_id = cur.lastrowid
        conn.commit()

        if deco_val or prod_val:
            dec_date = start.isoformat()
            target_loc = new_loc_id if is_base_mobile else loc_id
            if table_has_column("decorari", "rez_id"):
                cur.execute(
                    "INSERT INTO decorari (loc_id, rez_id, data, decor_cost, prod_cost, created_by) VALUES (?,?,?,?,?,?)",
                    (
                        target_loc,
                        rez_id,
                        dec_date,
                        deco_val,
                        prod_val,
                        user.get("username"),
                    ),
                )
            else:
                cur.execute(
                    "INSERT INTO decorari (loc_id, data, decor_cost, prod_cost, created_by) VALUES (?,?,?,?,?)",
                    (
                        target_loc,
                        dec_date,
                        deco_val,
                        prod_val,
                        user.get("username"),
                    ),
                )
            conn.commit()

        # actualizăm statusurile pe baza tuturor rezervărilor
        update_statusuri_din_rezervari()

        load_cb()
        win.destroy()

    ttk.Button(win, text="Confirmă închiriere", command=save_rent).grid(
        row=row_extra, column=0, columnspan=3, pady=10
    )


def open_edit_rent_window(root, rid, load_cb, parent=None):
    """Allow editing the rental period and price for reservation ``rid``.

    If *parent* is provided as ``(parent_id, old_start, old_end)``, the matching
    reservation for the parent location will be updated as well.
    """
    cur = conn.cursor()
    row = cur.execute(
        "SELECT data_start, data_end, suma FROM rezervari WHERE id=?", (rid,)
    ).fetchone()
    if not row:
        return

    ds, de, suma = row
    win = tk.Toplevel(root)
    win.title(f"Modifică închirierea #{rid}")

    ttk.Label(win, text="Data start:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    dp_start = DatePicker(win)
    dp_start.set_date(datetime.date.fromisoformat(ds))
    dp_start.grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(win, text="Data end:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    dp_end = DatePicker(win)
    dp_end.set_date(datetime.date.fromisoformat(de))
    dp_end.grid(row=1, column=1, padx=5, pady=5)

    ttk.Label(win, text="Sumă:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
    entry_fee = ttk.Entry(win, width=30)
    if suma is not None:
        entry_fee.insert(0, str(suma))
    entry_fee.grid(row=2, column=1, padx=5, pady=5)

    def save_edit():
        start = dp_start.get_date()
        end = dp_end.get_date()
        if start > end:
            messagebox.showwarning(
                "Interval incorect",
                "«Data start» trebuie înainte de «Data end».",
            )
            return

        fee_txt = entry_fee.get().strip()
        try:
            fee_val = float(fee_txt)
        except ValueError:
            messagebox.showwarning("Sumă invalidă", "Introdu o sumă numerică.")
            return

        cur.execute(
            "UPDATE rezervari SET data_start=?, data_end=?, suma=? WHERE id=?",
            (start.isoformat(), end.isoformat(), fee_val, rid),
        )
        if parent:
            pid, ods, ode = parent
            cur.execute(
                "UPDATE rezervari SET data_start=?, data_end=? WHERE loc_id=? AND data_start=? AND data_end=? AND suma=0",
                (start.isoformat(), end.isoformat(), pid, ods, ode),
            )
        conn.commit()
        update_statusuri_din_rezervari()
        load_cb()
        win.destroy()

    ttk.Button(win, text="Salvează", command=save_edit).grid(
        row=3, column=0, columnspan=2, pady=10
    )


def open_release_window(root, loc_id, load_cb, user):
    """Selectează și anulează o închiriere.

    Sunt listate doar închirierile active sau cele terminate în ultimele
    trei zile pentru a nu aglomera dialogul cu închirieri vechi.
    """
    cur = conn.cursor()
    cutoff = (datetime.date.today() - datetime.timedelta(days=3)).isoformat()
    rows = cur.execute(
        """
        SELECT id, client, data_start, data_end
          FROM rezervari
         WHERE loc_id=? AND suma IS NOT NULL AND data_end>=?
         ORDER BY data_start
        """,
        (loc_id, cutoff),
    ).fetchall()

    if not rows:
        messagebox.showinfo(
            "Eliberează", "Nu există închirieri pentru această locație."
        )
        return

    win = tk.Toplevel(root)
    win.title(f"Selectează închirierea #{loc_id}")

    lst = tk.Listbox(win, width=55, height=10)
    for rid, client, ds, de, *_ in rows:
        lst.insert("end", f"{client}: {ds} → {de}")
    lst.grid(row=0, column=0, columnspan=2, padx=5, pady=5)

    def delete_selected():
        sel = lst.curselection()
        if not sel:
            messagebox.showwarning("Selectează", "Alege o închiriere.")
            return
        rid, _client, ds, de, *_ = rows[sel[0]]

        end = datetime.date.fromisoformat(de)
        days_past = (datetime.date.today() - end).days
        if user.get("role") != "admin" and days_past > 3:
            messagebox.showwarning(
                "Refuzat",
                "Doar adminul poate șterge închirierile încheiate de mai mult de 3 zile.",
            )
            return

        if messagebox.askyesno(
            "Eliberează",
            "Ștergi complet perioada de închiriere?\nAlege 'Nu' pentru a modifica perioada.",
        ):
            cur.execute("DELETE FROM decorari WHERE rez_id=?", (rid,))
            cur.execute("DELETE FROM rezervari WHERE id=?", (rid,))
            parent_id = cur.execute(
                "SELECT parent_id FROM locatii WHERE id=?", (loc_id,)
            ).fetchone()
            if parent_id and parent_id[0]:
                cur.execute(
                    "DELETE FROM rezervari WHERE loc_id=? AND data_start=? AND data_end=? AND suma=0",
                    (parent_id[0], ds, de),
                )
                cur.execute("DELETE FROM locatii WHERE id=?", (loc_id,))
            conn.commit()
            update_statusuri_din_rezervari()
            load_cb()
            win.destroy()
        else:
            parent_id = cur.execute(
                "SELECT parent_id FROM locatii WHERE id=?", (loc_id,)
            ).fetchone()
            win.destroy()
            open_edit_rent_window(
                root,
                rid,
                load_cb,
                parent=(parent_id[0], ds, de) if parent_id and parent_id[0] else None,
            )

    ttk.Button(win, text="Confirmă", command=delete_selected).grid(
        row=1, column=0, padx=5, pady=5
    )
    ttk.Button(win, text="Închide", command=win.destroy).grid(
        row=1, column=1, padx=5, pady=5
    )


def open_decor_window(root, loc_id, user):
    """Add a decoration entry for a rented location."""
    loc_data = get_location_by_id(loc_id)
    win = tk.Toplevel(root)
    win.title("Adaugă decorare")

    ttk.Label(win, text="Data decorare:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    dp_date = DatePicker(win)
    dp_date.grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(win, text="Cost decorare:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    entry_deco = ttk.Entry(win, width=20)
    if loc_data and loc_data.get("decoration_cost"):
        entry_deco.insert(0, str(loc_data.get("decoration_cost")))
    entry_deco.grid(row=1, column=1, padx=5, pady=5)

    var_prod = tk.BooleanVar(value=False)
    chk_prod = ttk.Checkbutton(
        win,
        text="Include cost producție",
        variable=var_prod,
        command=lambda: entry_prod.grid() if var_prod.get() else entry_prod.grid_remove(),
    )
    chk_prod.grid(row=2, column=0, sticky="e", padx=5, pady=5)
    entry_prod = ttk.Entry(win, width=20)
    entry_prod.grid(row=2, column=1, padx=5, pady=5)
    entry_prod.grid_remove()

    def save():
        dec_date = dp_date.get_date().isoformat()
        try:
            dec_cost = float(entry_deco.get() or 0)
        except Exception:
            dec_cost = 0.0
        try:
            prod_cost = float(entry_prod.get() or 0) if var_prod.get() else 0.0
        except Exception:
            prod_cost = 0.0
        cur = conn.cursor()
        if table_has_column("decorari", "rez_id"):
            rez = cur.execute(
                "SELECT id FROM rezervari WHERE loc_id=? AND data_start<=? AND data_end>=? ORDER BY id DESC LIMIT 1",
                (loc_id, dec_date, dec_date),
            ).fetchone()
            rez_id = rez[0] if rez else None
            cur.execute(
                "INSERT INTO decorari (loc_id, rez_id, data, decor_cost, prod_cost, created_by) VALUES (?,?,?,?,?,?)",
                (loc_id, rez_id, dec_date, dec_cost, prod_cost, user.get("username")),
            )
        else:
            cur.execute(
                "INSERT INTO decorari (loc_id, data, decor_cost, prod_cost, created_by) VALUES (?,?,?,?,?)",
                (loc_id, dec_date, dec_cost, prod_cost, user.get("username")),
            )
        conn.commit()
        win.destroy()

    ttk.Button(win, text="Salvează", command=save).grid(row=3, column=0, columnspan=2, pady=10)


def open_edit_decor_window(root, decor_id, load_cb=None):
    """Edit the date and costs of a decoration entry."""
    cur = conn.cursor()
    row = cur.execute(
        "SELECT data, decor_cost, prod_cost FROM decorari WHERE id=?",
        (decor_id,),
    ).fetchone()
    if not row:
        return

    data_val, decor_val, prod_val = row

    win = tk.Toplevel(root)
    win.title(f"Modifică decorarea #{decor_id}")

    ttk.Label(win, text="Data decorare:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    dp_date = DatePicker(win)
    dp_date.set_date(datetime.date.fromisoformat(data_val))
    dp_date.grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(win, text="Cost decorare:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    entry_deco = ttk.Entry(win, width=20)
    if decor_val is not None:
        entry_deco.insert(0, str(decor_val))
    entry_deco.grid(row=1, column=1, padx=5, pady=5)

    ttk.Label(win, text="Cost producție:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
    entry_prod = ttk.Entry(win, width=20)
    if prod_val is not None:
        entry_prod.insert(0, str(prod_val))
    entry_prod.grid(row=2, column=1, padx=5, pady=5)

    def save():
        try:
            dec_cost = float(entry_deco.get() or 0)
        except Exception:
            dec_cost = 0.0
        try:
            prod_cost = float(entry_prod.get() or 0)
        except Exception:
            prod_cost = 0.0
        cur.execute(
            "UPDATE decorari SET data=?, decor_cost=?, prod_cost=? WHERE id=?",
            (dp_date.get_date().isoformat(), dec_cost, prod_cost, decor_id),
        )
        conn.commit()
        if load_cb:
            load_cb()
        win.destroy()

    ttk.Button(win, text="Salvează", command=save).grid(row=3, column=0, columnspan=2, pady=10)


def open_manage_decor_window(root, loc_id=None, load_cb=None):
    """Manage decoration entries for a location or all locations."""
    cur = conn.cursor()
    cutoff = (datetime.date.today() - datetime.timedelta(days=5)).isoformat()

    if loc_id:
        rows = cur.execute(
            "SELECT id, loc_id, data, decor_cost, prod_cost FROM decorari "
            "WHERE loc_id=? AND data>=? ORDER BY data",
            (loc_id, cutoff),
        ).fetchall()
    else:
        rows = []

    if not rows:
        rows = cur.execute(
            "SELECT id, loc_id, data, decor_cost, prod_cost FROM decorari "
            "WHERE data>=? ORDER BY data",
            (cutoff,),
        ).fetchall()
        loc_id = None

    if not rows:
        messagebox.showinfo("Decorări", "Nu există decorări de modificat.")
        return

    win = tk.Toplevel(root)

    if loc_id:
        loc = get_location_by_id(loc_id)
        if loc:
            win.title(f"Decorări: {loc.get('city')}, {loc.get('address')}")
            ttk.Label(
                win,
                text=f"{loc.get('city')}, {loc.get('address')}",
                font=("Segoe UI", 9, "bold"),
            ).grid(row=0, column=0, columnspan=3, padx=5, pady=(5, 0))
            row_start = 1
        else:
            win.title(f"Decorări locație #{loc_id}")
            row_start = 0
    else:
        win.title("Decorări recente")
        row_start = 0

    lst = tk.Listbox(win, width=50, height=10)
    for did, lid, data, dcost, pcost in rows:
        costs = []
        if dcost is not None:
            costs.append(str(dcost))
        if pcost is not None and pcost != 0:
            costs.append(f"prod {pcost}")
        cost_str = f" ({' + '.join(costs)})" if costs else ""
        if loc_id is None:
            loc = get_location_by_id(lid)
            loc_info = f"{loc.get('city')}, {loc.get('address')}" if loc else f"#{lid}"
            lst.insert("end", f"{loc_info}: {data}{cost_str}")
        else:
            lst.insert("end", f"{data}{cost_str}")
    lst.grid(row=row_start, column=0, columnspan=3, padx=5, pady=5)

    def delete_selected():
        sel = lst.curselection()
        if not sel:
            messagebox.showwarning("Selectează", "Alege o decorare.")
            return
        did, _lid, data, *_ = rows[sel[0]]
        if messagebox.askyesno("Șterge", "Ștergi decorarea selectată?"):
            cur.execute("DELETE FROM decorari WHERE id=?", (did,))
            conn.commit()
            win.destroy()
            load_cb()

    def edit_selected():
        sel = lst.curselection()
        if not sel:
            messagebox.showwarning("Selectează", "Alege o decorare.")
            return
        did, *_ = rows[sel[0]]
        open_edit_decor_window(win, did, load_cb)

    btn_row = row_start + 1
    ttk.Button(win, text="Editează", command=edit_selected).grid(row=btn_row, column=0, padx=5, pady=5)
    ttk.Button(win, text="Șterge", command=delete_selected).grid(row=btn_row, column=1, padx=5, pady=5)
    ttk.Button(win, text="Închide", command=win.destroy).grid(row=btn_row, column=2, padx=5, pady=5)


def export_available_excel(
    grup_filter, status_filter, search_term, ignore_dates, start_date, end_date
):
    import datetime
    import pandas as pd
    from tkinter import messagebox, filedialog
    from db import read_sql_query

    # 1) Construim WHERE identic cu load_locations()
    cond, params = [], []
    if grup_filter and grup_filter != "Toate":
        cond.append("grup = ?")
        params.append(grup_filter)
    if status_filter and status_filter != "Toate":
        cond.append("status = ?")
        params.append(status_filter)
    if search_term:
        cond.append("(city LIKE ? OR county LIKE ? OR address LIKE ?)")
        params += [f"%{search_term}%"] * 3
    if not ignore_dates:
        cond.append(
            """
            (data_start IS NULL
             OR data_end   IS NULL
             OR data_end   < ?
             OR data_start > ?)
        """
        )
        params += [start_date.isoformat(), end_date.isoformat()]

    where = ("WHERE " + " AND ".join(cond)) if cond else ""
    sql = f"""
        SELECT
          city, county, address, type,
          gps, photo_link,
          size, sqm, illumination,
          ratecard, decoration_cost,
          data_start, data_end,
          grup, status
        FROM locatii
        {where}
        ORDER BY grup, city, id
    """

    # 2) Citim datele
    df = read_sql_query(sql, params=params, parse_dates=["data_start", "data_end"])
    if df.empty:
        messagebox.showinfo(
            "Export Excel", "Nu există locații pentru criteriile alese."
        )
        return

    # 3) Calculăm mesajul de Availability
    today = datetime.datetime.now().date()

    def avail_msg(r):
        ds, de = r["data_start"], r["data_end"]
        if pd.notna(ds) and ds.date() > today:
            until = (ds.date() - datetime.timedelta(days=1)).strftime("%d.%m.%Y")
            return f"Disponibil până la {until}"
        if pd.notna(de) and de.date() >= today:
            frm = (de.date() + datetime.timedelta(days=1)).strftime("%d.%m.%Y")
            return f"Disponibil din {frm}"
        if r["status"] != "Disponibil" and pd.notna(de) and de.date() < today:
            frm = (de.date() + datetime.timedelta(days=1)).strftime("%d.%m.%Y")
            return f"Disponibil din {frm}"
        return "Disponibil"

    df["Availability"] = df.apply(avail_msg, axis=1)

    # 4) Coloane de export
    write_cols = [
        "city",
        "county",
        "address",
        "type",
        "gps",
        "photo_link",
        "size",
        "sqm",
        "illumination",
        "ratecard",
        "decoration_cost",
        "Availability",
    ]

    # 5) Alege fișierul de salvat
    fp = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Salvează fișierul Excel",
    )
    if not fp:
        return

    # 6) Scriem Excel: câte o foaie per grup
    with pd.ExcelWriter(fp, engine="xlsxwriter") as writer:
        wb = writer.book
        center_fmt = wb.add_format(
            {"align": "center", "valign": "vcenter", "border": 1}
        )
        money_fmt = wb.add_format(
            {
                "num_format": "€#,##0.00",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )
        link_fmt = wb.add_format(
            {
                "font_color": "blue",
                "underline": True,
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )
        title_fmt = wb.add_format(
            {"bold": True, "font_size": 14, "align": "center", "valign": "vcenter"}
        )
        hdr_fmt = wb.add_format(
            {
                "bold": True,
                "bg_color": "#4F81BD",
                "font_color": "white",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )

        for grup, sub in df.groupby("grup"):
            grp_name = (grup or "").strip() or "FaraGrup"
            sheet = grp_name[:31]
            title = f"Locații {grp_name}"

            sub_df = sub.loc[:, write_cols].copy()
            sub_df.columns = [
                "City",
                "County",
                "Address",
                "Type",
                "GPS",
                "Photo Link",
                "Size",
                "SQM",
                "Illum",
                "Rate Card",
                "Installation & Removal",
                "Availability",
            ]
            sub_df.insert(0, "Nr", range(1, len(sub_df) + 1))

            startrow = 1
            ws = writer.book.add_worksheet(sheet)
            writer.sheets[sheet] = ws

            last_col = len(sub_df.columns) - 1
            ws.merge_range(0, 0, 0, last_col, title, title_fmt)

            for col_idx, name in enumerate(sub_df.columns):
                ws.write(startrow, col_idx, name, hdr_fmt)

            for row_idx, row in enumerate(
                sub_df.itertuples(index=False), start=startrow + 1
            ):
                for col_idx, value in enumerate(row):
                    col_name = sub_df.columns[col_idx]
                    if pd.isna(value):
                        value = ""
                    if col_name in ("Rate Card", "Installation & Removal"):
                        fmt = money_fmt
                    else:
                        fmt = center_fmt
                    ws.write(row_idx, col_idx, value, fmt)

            for idx, col_name in enumerate(sub_df.columns):
                if col_name in ("Rate Card", "Installation & Removal"):
                    vals = pd.to_numeric(sub_df[col_name], errors="coerce").fillna(0)
                    formatted = [f"€{v:,.2f}" for v in vals]
                    max_len = max(len(col_name), *(len(v) for v in formatted))
                elif col_name == "GPS":
                    max_len = max(len(col_name), len("Maps"))
                elif col_name == "Photo Link":
                    max_len = max(len(col_name), len("Photo"))
                else:
                    max_len = max(
                        len(col_name), sub_df[col_name].astype(str).map(len).max()
                    )
                ws.set_column(idx, idx, max_len + 2)

            gi = sub_df.columns.get_loc("GPS")
            for r, coord in enumerate(sub["gps"], start=startrow + 1):
                if coord:
                    url = f"https://www.google.com/maps/search/?api=1&query={coord}"
                    ws.write_url(r, gi, url, link_fmt, string="Maps")

            pi = sub_df.columns.get_loc("Photo Link")
            for r, u in enumerate(sub["photo_link"], start=startrow + 1):
                if u and u.strip():
                    url = u.strip()
                    if not url.lower().startswith(("http://", "https://")):
                        url = "https://" + url
                    ws.write_url(r, pi, url, link_fmt, string="Photo")
                else:
                    ws.write(r, pi, "", center_fmt)

    messagebox.showinfo("Export Excel", f"Am salvat locațiile în:\n{fp}")


def export_sales_report():
    """Exportă un raport structurat pe luni cu informații despre vânzări."""
    import pandas as pd
    import datetime
    from tkinter import messagebox, filedialog
    from db import read_sql_query, update_statusuri_din_rezervari

    update_statusuri_din_rezervari()

    year = choose_report_year()
    if year is None:
        return

    df_loc_all = read_sql_query(
        """
        SELECT id, city, county, address, type, size, sqm, illumination,
               ratecard, pret_vanzare, grup, status, is_mobile, parent_id
          FROM locatii
         ORDER BY county, city, id
        """,
    )
    # ``df_loc`` is used for the monthly sheets where the base record for mobile
    # prisms would only duplicate the actual rented units.  Keep the full
    # dataframe separate so the "Total" sheet can include mobile bases as well.
    df_loc = df_loc_all[
        ~((df_loc_all["is_mobile"] == 1) & (df_loc_all["parent_id"].isna()))
    ]

    df_total_base = df_loc_all[df_loc_all["parent_id"].isna()].copy()

    if df_loc.empty:
        messagebox.showinfo("Export Excel", "Nu există locații în baza de date.")
        return

    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        title="Salvează raportul Excel",
    )
    if not path:
        return

    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        wb = writer.book
        stats_ranges = {}
        hdr_fmt = wb.add_format(
            {
                "bold": True,
                "bg_color": "#4F81BD",
                "font_color": "white",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )
        text_fmt = wb.add_format({"align": "center", "valign": "vcenter", "border": 1})
        money_fmt = wb.add_format(
            {
                "num_format": "€#,##0.00",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )
        sold_text_fmt = wb.add_format(
            {"align": "center", "valign": "vcenter", "border": 1, "bg_color": "#D9E1F2"}
        )
        sold_money_fmt = wb.add_format(
            {
                "num_format": "€#,##0.00",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
                "bg_color": "#D9E1F2",
            }
        )
        percent_fmt = wb.add_format(
            {"num_format": "0.00%", "align": "center", "valign": "vcenter", "border": 1}
        )

        stat_lbl_fmt = wb.add_format(
            {
                "bold": True,
                "bg_color": "white",
                "font_size": 14,
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )
        stat_money_fmt = wb.add_format(
            {
                "num_format": "€#,##0.00",
                "bold": True,
                "bg_color": "white",
                "font_size": 14,
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )
        stat_money_pos_fmt = wb.add_format(
            {
                "num_format": "€#,##0.00",
                "bold": True,
                "bg_color": "white",
                "font_size": 14,
                "font_color": "green",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )
        stat_money_neg_fmt = wb.add_format(
            {
                "num_format": "€#,##0.00",
                "bold": True,
                "bg_color": "white",
                "font_size": 14,
                "font_color": "red",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )
        stat_percent_fmt = wb.add_format(
            {
                "num_format": "0.00%",
                "bold": True,
                "bg_color": "white",
                "font_size": 14,
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )
        stat_int_fmt = wb.add_format(
            {
                "bold": True,
                "bg_color": "white",
                "font_size": 14,
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )

        money_cols = {
            "Ratecard/month",
            "PRET DE VANZARE",
            "PRET DE INCHIRIERE",
            "SUMĂ AN",
        }

        def write_sheet(name, df_sheet):
            df_sheet = df_sheet.copy()
            df_sheet["data_start"] = pd.to_datetime(df_sheet["data_start"], errors="coerce")
            df_sheet["data_end"] = pd.to_datetime(df_sheet["data_end"], errors="coerce")
            df_sheet["Perioada"] = ""
            mask = df_sheet["data_start"].notna() & df_sheet["data_end"].notna()
            if mask.any():
                df_sheet.loc[mask, "Perioada"] = (
                    df_sheet.loc[mask, "data_start"].dt.strftime("%d.%m.%Y")
                    + " → "
                    + df_sheet.loc[mask, "data_end"].dt.strftime("%d.%m.%Y")
                )
            df_sheet = df_sheet[
                [
                    "city",
                    "county",
                    "address",
                    "type",
                    "size",
                    "sqm",
                    "illumination",
                    "ratecard",
                    "pret_vanzare",
                    "suma",
                    "client",
                    "Perioada",
                    "status",
                ]
            ]
            df_sheet.columns = [
                "City",
                "County",
                "Address",
                "Type",
                "Size",
                "SQM",
                "Illum",
                "Ratecard/month",
                "PRET DE VANZARE",
                "PRET DE INCHIRIERE",
                "Client",
                "Perioada",
                "status",
            ]
            df_sheet.insert(0, "Nr", range(1, len(df_sheet) + 1))

            ws = wb.add_worksheet(name)
            writer.sheets[name] = ws

            for col_idx, col_name in enumerate(df_sheet.columns[:-1]):
                ws.write(0, col_idx, col_name, hdr_fmt)

            for row_idx, row in enumerate(df_sheet.itertuples(index=False), start=1):
                sold = row.status == "Închiriat"
                for col_idx, value in enumerate(row[:-1]):
                    col_name = df_sheet.columns[col_idx]
                    if pd.isna(value):
                        value = ""
                    if col_name in money_cols:
                        fmt = sold_money_fmt if sold else money_fmt
                    else:
                        fmt = sold_text_fmt if sold else text_fmt
                    ws.write(row_idx, col_idx, value, fmt)

            for idx, col_name in enumerate(df_sheet.columns[:-1]):
                if col_name in money_cols:
                    vals = pd.to_numeric(df_sheet[col_name], errors="coerce").fillna(0)
                    formatted = [f"€{v:,.2f}" for v in vals]
                    max_len = max(len(col_name), *(len(v) for v in formatted))
                else:
                    max_len = max(
                        len(col_name), df_sheet[col_name].astype(str).map(len).max()
                    )
                ws.set_column(idx, idx, max_len + 2)

            sold_mask = df_sheet["status"] == "Închiriat"
            pct_sold = sold_mask.mean()
            pct_free = 1 - pct_sold
            count_sold = int(sold_mask.sum())
            count_free = int(len(df_sheet) - count_sold)
            sale_total = (
                pd.to_numeric(df_sheet["PRET DE VANZARE"], errors="coerce")
                .fillna(0)
                .sum()
            )
            sold_income = (
                pd.to_numeric(
                    df_sheet.loc[sold_mask, "PRET DE INCHIRIERE"], errors="coerce"
                )
                .fillna(0)
                .sum()
            )
            # Value of the unsold locations based on their sale price
            sale_free = (
                pd.to_numeric(
                    df_sheet.loc[~sold_mask, "PRET DE VANZARE"], errors="coerce"
                )
                .fillna(0)
                .sum()
            )

            # shift the monthly totals block one column to the left by
            # ignoring the final "status" column when computing the end index
            stats_end = len(df_sheet.columns) - 2
            value_col = max(stats_end - 2, 0)
            merge_end = value_col - 1
            merge_start = max(merge_end - 3, 0)
            start = len(df_sheet) + 2
            ws.merge_range(start, merge_start, start, merge_end, "Locații vândute", stat_lbl_fmt)
            ws.merge_range(
                start,
                value_col,
                start,
                stats_end,
                f"{count_sold} ({pct_sold:.2%})",
                stat_int_fmt,
            )
            ws.merge_range(
                start + 1, merge_start, start + 1, merge_end, "Locații nevândute", stat_lbl_fmt
            )
            ws.merge_range(
                start + 1,
                value_col,
                start + 1,
                stats_end,
                f"{count_free} ({pct_free:.2%})",
                stat_int_fmt,
            )
            ws.merge_range(
                start + 2, merge_start, start + 2, merge_end, "Preț vânzare total", stat_lbl_fmt
            )
            ws.merge_range(
                start + 2,
                value_col,
                start + 2,
                stats_end,
                sale_total,
                stat_money_fmt,
            )
            pct_sale_sold = sold_income / sale_total if sale_total else 0
            pct_sale_free = sale_free / sale_total if sale_total else 0
            ws.merge_range(
                start + 3, merge_start, start + 3, merge_end, "Sumă locații vândute", stat_lbl_fmt
            )
            ws.merge_range(
                start + 3,
                value_col,
                start + 3,
                stats_end,
                f"€{sold_income:,.2f} ({pct_sale_sold:.2%})",
                stat_money_pos_fmt,
            )
            ws.merge_range(
                start + 4,
                merge_start,
                start + 4,
                merge_end,
                "Sumă locații nevândute",
                stat_lbl_fmt,
            )
            ws.merge_range(
                start + 4,
                value_col,
                start + 4,
                stats_end,
                f"€{sale_free:,.2f} ({pct_sale_free:.2%})",
                stat_money_neg_fmt,
            )
            stats_ranges[name] = (start, start + 4, merge_start, stats_end)
            # The "Raport sume vândute/nevândute" statistic is no longer shown
            # in the monthly sheets as it was not considered relevant.

        year_start = datetime.date(year, 1, 1)
        year_end = datetime.date(year, 12, 31)
        df_rez = read_sql_query(
            """
            SELECT l.id, l.grup, l.city, l.county, l.address, l.type, l.size, l.sqm, l.illumination,
                   l.ratecard, l.pret_vanzare, l.is_mobile, l.parent_id,
                   r.client, r.data_start, r.data_end, r.suma
              FROM rezervari r
              JOIN locatii l ON r.loc_id = l.id
             WHERE r.suma IS NOT NULL AND r.suma > 0
               AND NOT (r.data_end < ? OR r.data_start > ?)
             ORDER BY r.data_start
            """,
            params=[year_start.isoformat(), year_end.isoformat()],
            parse_dates=["data_start", "data_end"],
        )
        df_rez = df_rez[~((df_rez["is_mobile"] == 1) & (df_rez["parent_id"].isna()))]

        df_all = df_loc[
            [
                "id",
                "city",
                "county",
                "address",
                "type",
                "size",
                "sqm",
                "illumination",
                "ratecard",
                "pret_vanzare",
                "grup",
                "status",
                "is_mobile",
                "parent_id",
            ]
        ].copy()
        df_base = df_total_base.copy()

        # Calculăm numărul total de zile vândute și valoarea reală pentru fiecare locație
        import calendar

        records = []
        for row in df_rez.itertuples(index=False):
            base_id = row.parent_id if pd.notna(row.parent_id) else row.id
            start = max(row.data_start.date(), year_start)
            end = min(row.data_end.date(), year_end)
            cur = start
            while cur <= end:
                dim = calendar.monthrange(cur.year, cur.month)[1]
                month_end = datetime.date(cur.year, cur.month, dim)
                ov_end = min(end, month_end)
                days = (ov_end - cur).days + 1
                frac = days / dim
                records.append(
                    {
                        "id": base_id,
                        "days": days,
                        "months": frac,
                        "val_real": row.suma * frac,
                    }
                )
                cur = ov_end + datetime.timedelta(days=1)

        df_rec = pd.DataFrame(records)
        if df_rec.empty:
            df_rec = pd.DataFrame(columns=["id", "days", "months", "val_real"])
        agg = df_rec.groupby("id").agg(
            {"days": "sum", "months": "sum", "val_real": "sum"}
        )

        parent_ids = df_rez["parent_id"].fillna(df_rez["id"]).astype("Int64")
        units_sold = df_rez.groupby(parent_ids)["id"].nunique()

        # Sheet summarizing the entire year
        df_total = df_base.copy()
        df_total["Sold Days"] = df_total["id"].map(agg["days"]).fillna(0)
        df_total["Sold Months"] = df_total["id"].map(agg["months"]).fillna(0)
        df_total["Units Sold"] = (
            df_total["id"].map(units_sold).fillna(0).astype(int)
        )
        df_total["pret_vanzare"] = pd.to_numeric(
            df_total["pret_vanzare"], errors="coerce"
        ).fillna(0)
        df_total["Total Sum"] = df_total["id"].map(agg["val_real"]).fillna(0)
        mob_mask = df_total["is_mobile"] == 1
        df_total.loc[mob_mask & (df_total["Units Sold"] > 0), "address"] = (
            df_total.loc[mob_mask & (df_total["Units Sold"] > 0), "address"]
            + " "
            + df_total.loc[mob_mask & (df_total["Units Sold"] > 0), "Units Sold"].astype(str)
            + "X"
        )
        df_total = df_total.drop(columns="Units Sold")
        df_total["% Year Sold"] = df_total.apply(
            lambda r: (
                r["Sold Months"] / (20 * 12)
                if r["is_mobile"]
                else (r["Sold Months"] / 12)
            ),
            axis=1,
        )
        grp_order = (
            df_total.groupby("grup")["pret_vanzare"]
            .max()
            .sort_values(ascending=False)
            .index
        )
        order_map = {g: i for i, g in enumerate(grp_order)}
        df_total["__grp"] = df_total["grup"].map(order_map)
        df_total = df_total.sort_values(
            ["__grp", "pret_vanzare"], ascending=[True, False]
        ).drop(columns="__grp")

        def write_total_sheet(df_sheet):
            df_sheet = df_sheet[
                [
                    "city",
                    "county",
                    "address",
                    "type",
                    "size",
                    "sqm",
                    "illumination",
                    "ratecard",
                    "pret_vanzare",
                    "Sold Months",
                    "% Year Sold",
                    "Total Sum",
                ]
            ].copy()
            df_sheet.columns = [
                "City",
                "County",
                "Address",
                "Type",
                "Size",
                "SQM",
                "Illum",
                "Ratecard/month",
                "PRET DE VANZARE",
                "Luni vândută",
                "% An vândut",
                "SUMĂ AN",
            ]
            df_sheet.insert(0, "Nr", range(1, len(df_sheet) + 1))
            ws = wb.add_worksheet("Total")
            writer.sheets["Total"] = ws
            for col_idx, col_name in enumerate(df_sheet.columns):
                ws.write(0, col_idx, col_name, hdr_fmt)
            for row_idx, row in enumerate(df_sheet.itertuples(index=False), start=1):
                for col_idx, value in enumerate(row):
                    col_name = df_sheet.columns[col_idx]
                    if col_name in {"PRET DE VANZARE", "SUMĂ AN"}:
                        fmt = money_fmt
                    else:
                        fmt = text_fmt
                    if col_name == "% An vândut":
                        ws.write(row_idx, col_idx, value, percent_fmt)
                    else:
                        ws.write(row_idx, col_idx, value, fmt)
            for idx, col in enumerate(df_sheet.columns):
                if col in {"PRET DE VANZARE", "SUMĂ AN"}:
                    vals = pd.to_numeric(df_sheet[col], errors="coerce").fillna(0)
                    formatted = [f"€{v:,.2f}" for v in vals]
                    width = max(len(col), *(len(v) for v in formatted)) + 2
                else:
                    width = max(len(col), df_sheet[col].astype(str).map(len).max()) + 2
                ws.set_column(idx, idx, width)
            pct_months_sold = (
                pd.to_numeric(df_sheet["Luni vândută"], errors="coerce").fillna(0).sum()
                / (len(df_sheet) * 12)
            )
            pct_months_free = 1 - pct_months_sold
            sale_total = (
                pd.to_numeric(df_sheet["PRET DE VANZARE"], errors="coerce").fillna(0) * 12
            ).sum()
            sold_income = (
                pd.to_numeric(df_sheet["SUMĂ AN"], errors="coerce").fillna(0).sum()
            )
            sale_free = sale_total - sold_income
            pct_sale_sold = sold_income / sale_total if sale_total else 0
            pct_sale_free = sale_free / sale_total if sale_total else 0

            stats_end = len(df_sheet.columns) - 1
            value_col = max(stats_end - 2, 0)
            merge_end = value_col - 1
            merge_start = max(merge_end - 3, 0)
            start = len(df_sheet) + 2
            ws.merge_range(
                start,
                merge_start,
                start,
                merge_end,
                f"Locații vândute în anul {year}",
                stat_lbl_fmt,
            )
            ws.merge_range(
                start,
                value_col,
                start,
                stats_end,
                pct_months_sold,
                stat_percent_fmt,
            )
            ws.merge_range(
                start + 1, merge_start, start + 1, merge_end, "Locații nevândute", stat_lbl_fmt
            )
            ws.merge_range(
                start + 1,
                value_col,
                start + 1,
                stats_end,
                pct_months_free,
                stat_percent_fmt,
            )
            ws.merge_range(
                start + 2, merge_start, start + 2, merge_end, "Preț vânzare total", stat_lbl_fmt
            )
            ws.merge_range(
                start + 2,
                value_col,
                start + 2,
                stats_end,
                sale_total,
                stat_money_fmt,
            )
            ws.merge_range(
                start + 3, merge_start, start + 3, merge_end, "Sumă locații vândute", stat_lbl_fmt
            )
            ws.merge_range(
                start + 3,
                value_col,
                start + 3,
                stats_end,
                f"€{sold_income:,.2f} ({pct_sale_sold:.2%})",
                stat_money_pos_fmt,
            )
            ws.merge_range(
                start + 4, merge_start, start + 4, merge_end, "Sumă locații nevândute", stat_lbl_fmt
            )
            ws.merge_range(
                start + 4,
                value_col,
                start + 4,
                stats_end,
                f"€{sale_free:,.2f} ({pct_sale_free:.2%})",
                stat_money_neg_fmt,
            )
            stats_ranges["Total"] = (start, start + 4, merge_start, stats_end)

        write_total_sheet(df_total)

        for month in range(1, 13):
            start_m = pd.Timestamp(year, month, 1)
            end_m = start_m + pd.offsets.MonthEnd(0)
            mask = (df_rez["data_end"] >= start_m) & (df_rez["data_start"] <= end_m)
            sub = df_rez.loc[mask].copy()
            if not sub.empty:
                mdays = calendar.monthrange(start_m.year, start_m.month)[1]
                sub["overlap_start"] = sub["data_start"].clip(lower=start_m)
                sub["overlap_end"] = sub["data_end"].clip(upper=end_m)
                sub["suma"] = sub.apply(
                    lambda r: r.suma
                    * ((r.overlap_end - r.overlap_start).days + 1)
                    / mdays,
                    axis=1,
                )
                grouped = (
                    sub.groupby("id")
                    .agg(
                        {
                            "client": "first",
                            "overlap_start": "min",
                            "overlap_end": "max",
                            "suma": "sum",
                        }
                    )
                    .reset_index()
                    .rename(
                        columns={
                            "overlap_start": "data_start",
                            "overlap_end": "data_end",
                        }
                    )
                )
            else:
                grouped = pd.DataFrame(
                    columns=["id", "client", "data_start", "data_end", "suma"]
                )
            df_month = df_all.merge(grouped, on="id", how="left")
            df_month["status"] = df_month["client"].apply(
                lambda x: "Închiriat" if pd.notna(x) else "Disponibil"
            )
            df_month["pret_vanzare"] = pd.to_numeric(
                df_month["pret_vanzare"], errors="coerce"
            ).fillna(0)
            grp_order = (
                df_month.groupby("grup")["pret_vanzare"]
                .max()
                .sort_values(ascending=False)
                .index
            )
            order_map = {g: i for i, g in enumerate(grp_order)}
            df_month["__grp"] = df_month["grup"].map(order_map)
            df_month = df_month.sort_values(
                ["__grp", "pret_vanzare"], ascending=[True, False]
            ).drop(columns="__grp")
            name = start_m.strftime("%B")
            write_sheet(name, df_month)

    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side

    wb = load_workbook(path)
    thin = Side(style="thin")
    thick = Side(style="thick")
    for sheet_name, (r1, r2, c1, c2) in stats_ranges.items():
        ws = wb[sheet_name]
        start_row, end_row = r1 + 1, r2 + 1
        start_col, end_col = c1 + 1, c2 + 1
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                border = cell.border
                left = thick if c == start_col else border.left
                right = thick if c == end_col else border.right
                top = thick if r == start_row else border.top
                bottom = thick if r == end_row else border.bottom
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)
    wb.save(path)

    messagebox.showinfo("Export Excel", f"Raport salvat:\n{path}")


def export_decor_report():
    """Export a simple Excel report with all decorations for a year."""
    import pandas as pd
    from tkinter import filedialog, messagebox
    from db import read_sql_query

    year = choose_report_year()
    if year is None:
        return

    start = f"{year}-01-01"
    end = f"{year}-12-31"
    df = read_sql_query(
        """
        SELECT d.data, l.city, l.county, l.address, l.code,
               d.decor_cost, d.prod_cost, d.created_by
          FROM decorari d
          JOIN locatii l ON d.loc_id = l.id
         WHERE d.data BETWEEN ? AND ?
         ORDER BY d.data
        """,
        params=[start, end],
        parse_dates=["data"],
    )

    if df.empty:
        messagebox.showinfo("Raport", "Nu există decorări pentru anul selectat.")
        return

    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        title="Salvează raportul",
    )
    if not path:
        return

    df.rename(
        columns={
            "data": "Data",
            "city": "Oraș",
            "county": "Județ",
            "address": "Adresă",
            "code": "Cod",
            "decor_cost": "Cost Decor",
            "prod_cost": "Cost Producție",
            "created_by": "Adăugat de",
        },
        inplace=True,
    )
    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Decorări")

    messagebox.showinfo("Raport", f"Raport salvat:\n{path}")


def open_offer_window(tree):
    import pandas as pd
    import datetime
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
    from db import read_sql_query

    # 1. Preluare selecție
    sel = tree.selection()
    if not sel:
        messagebox.showwarning("Nimic selectat", "Selectează cel puțin o locație.")
        return
    ids = [int(i) for i in sel]

    # 2. Fereastră setări ofertă
    win = tk.Toplevel(tree.master)
    win.title("Setări ofertă personalizată")

    # 2a. Parametri generali
    params = [
        ("Discount (%)", "20"),
        ("Decorare / m² (€)", "10"),
        ("Producție / m² (€)", "5"),
    ]
    entries = {}
    for i, (label, default) in enumerate(params):
        ttk.Label(win, text=label + ":").grid(
            row=i, column=0, sticky="e", padx=5, pady=3
        )
        e = ttk.Entry(win, width=10)
        e.insert(0, default)
        e.grid(row=i, column=1, padx=5, pady=3)
        entries[label] = e

    # 2b. Alegere Preț de Bază
    price_var = tk.StringVar(value="ratecard")
    ttk.Label(win, text="Preț de bază:").grid(
        row=3, column=0, sticky="e", padx=5, pady=3
    )
    frm_price = ttk.Frame(win)
    frm_price.grid(row=3, column=1, sticky="w", padx=5, pady=3)
    ttk.Radiobutton(
        frm_price, text="Rate Card", variable=price_var, value="ratecard"
    ).pack(side="left")
    ttk.Radiobutton(
        frm_price, text="Preț Vânzare", variable=price_var, value="pret_vanzare"
    ).pack(side="left")

    # 2c. Checkbox Discount personalizat
    personal_var = tk.BooleanVar(value=False)
    chk_personal = ttk.Checkbutton(
        win, text="Discount personalizat per locație", variable=personal_var
    )
    chk_personal.grid(row=4, column=0, columnspan=2, pady=5)

    def export():
        # 3. Validare input
        try:
            disc_pct = float(entries["Discount (%)"].get())
            cost_deco = float(entries["Decorare / m² (€)"].get())
            cost_prod = float(entries["Producție / m² (€)"].get())
        except ValueError:
            messagebox.showerror("Date invalide", "Introdu valori numerice valide.")
            return

        # 4. Citire date din DB (adăugăm address + pret_vanzare)
        sql = (
            f"SELECT id, city, county, address, gps, code, photo_link, sqm, type, "
            f"ratecard, pret_vanzare, data_start, data_end, is_mobile, parent_id "
            f"FROM locatii WHERE id IN ({','.join(['?']*len(ids))})"
        )
        df = read_sql_query(
            sql,
            params=ids,
            parse_dates=["data_start", "data_end"],
        )

        # 4a. Cantitatea pentru locațiile mobile de bază
        df["qty"] = 1
        for idx, row in df.iterrows():
            if row.get("is_mobile") and not row.get("parent_id"):
                q = simpledialog.askinteger(
                    "Cantitate",
                    f"Câte bucăți pentru {row['address']}?",
                    parent=win,
                    minvalue=1,
                    maxvalue=20,
                    initialvalue=1,
                )
                if q:
                    df.at[idx, "qty"] = int(q)
                    df.at[idx, "address"] = f"{row['address']} * {int(q)}"

        # 5. Calcul disponibilitate
        today = datetime.date.today()

        def avail(r):
            if r.get("is_mobile") and not r.get("parent_id"):
                return "Disponibil"
            ds, de = r["data_start"], r["data_end"]
            if pd.notna(ds) and ds.date() > today:
                return f"Până pe {(ds.date() - datetime.timedelta(days=1)).strftime('%d.%m.%Y')}"
            if pd.notna(de) and de.date() >= today:
                return f"Din {(de.date() + datetime.timedelta(days=1)).strftime('%d.%m.%Y')}"
            return "Disponibil"

        df["Availability"] = df.apply(avail, axis=1)

        # 6. Calcul costuri
        df["Installation & Removal"] = df["sqm"] * cost_deco * df["qty"]
        df["Production"] = df["sqm"] * cost_prod * df["qty"]

        # 7. Alege prețul de bază
        base_col = price_var.get()  # 'ratecard' sau 'pret_vanzare'
        df["Base Price"] = df[base_col].fillna(0).astype(float) * df["qty"]

        # Funcție comună de scriere Excel
        def write_excel(df_export):
            df_export = df_export.sort_values(by=["City", "CODE"])
            df_export.insert(0, "No.", range(1, len(df_export) + 1))

            fp = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Salvează oferta",
            )
            if not fp:
                return

            with pd.ExcelWriter(fp, engine="xlsxwriter") as writer:
                sheet = "Ofertă"
                wb = writer.book
                ws = wb.add_worksheet(sheet)
                writer.sheets[sheet] = ws

                title_fmt = wb.add_format(
                    {
                        "align": "center",
                        "valign": "vcenter",
                        "bold": True,
                        "font_size": 14,
                    }
                )
                hdr_fmt = wb.add_format(
                    {
                        "align": "center",
                        "valign": "vcenter",
                        "bold": True,
                        "bg_color": "#4F81BD",
                        "font_color": "white",
                        "border": 1,
                    }
                )
                txt_fmt = wb.add_format(
                    {"align": "center", "valign": "vcenter", "border": 1}
                )
                money_fmt = wb.add_format(
                    {
                        "align": "center",
                        "valign": "vcenter",
                        "num_format": "€#,##0.00",
                        "border": 1,
                    }
                )
                link_fmt = wb.add_format(
                    {
                        "font_color": "blue",
                        "underline": True,
                        "align": "center",
                        "valign": "vcenter",
                        "border": 1,
                    }
                )
                percent_fmt = wb.add_format(
                    {
                        "align": "center",
                        "valign": "vcenter",
                        "num_format": "0.00%",
                        "border": 1,
                    }
                )

                last_col = len(df_export.columns) - 1
                ws.merge_range(0, 0, 0, last_col, "OFERTĂ PERSONALIZATĂ", title_fmt)

                for col_idx, hdr in enumerate(df_export.columns):
                    ws.write(1, col_idx, hdr, hdr_fmt)

                money_cols = {
                    "Base Price",
                    "Final Price",
                    "Installation & Removal",
                    "Production",
                }
                percent_cols = {"% Discount"}

                for row_idx, row in enumerate(
                    df_export.itertuples(index=False), start=2
                ):
                    for col_idx, value in enumerate(row):
                        col_name = df_export.columns[col_idx]
                        if col_name in money_cols:
                            fmt = money_fmt
                        elif col_name in percent_cols:
                            fmt = percent_fmt
                        else:
                            fmt = txt_fmt
                        ws.write(row_idx, col_idx, value, fmt)

                for idx, col in enumerate(df_export.columns):
                    if col in money_cols:
                        vals = pd.to_numeric(df_export[col], errors="coerce").fillna(0)
                        width = max(len(col), *(len(f"€{v:,.2f}") for v in vals)) + 2
                    elif col == "GPS":
                        width = max(len(col), len("Maps")) + 2
                    elif col == "Photo Link":
                        width = max(len(col), len("Photo")) + 2
                    else:
                        width = (
                            max(len(col), df_export[col].astype(str).map(len).max()) + 2
                        )
                    ws.set_column(idx, idx, width)

                gps_idx = df_export.columns.get_loc("GPS")
                for r, coord in enumerate(df_export["GPS"], start=2):
                    if isinstance(coord, str) and coord.strip():
                        ws.write_url(
                            r,
                            gps_idx,
                            f"https://www.google.com/maps/search/?api=1&query={coord.strip()}",
                            link_fmt,
                            string="Maps",
                        )

                photo_idx = df_export.columns.get_loc("Photo Link")
                for r, url in enumerate(df_export["Photo Link"], start=2):
                    if isinstance(url, str) and url.strip():
                        link = url.strip()
                        if not link.lower().startswith(("http://", "https://")):
                            link = "https://" + link
                        ws.write_url(r, photo_idx, link, link_fmt, string="Photo")

            messagebox.showinfo("Succes", f"Fișierul a fost salvat:\n{fp}")
            win.destroy()

        # 8. Branch pentru discount personalizat
        if personal_var.get():
            win2 = tk.Toplevel(win)
            win2.title("Discount pe locație")
            entry_map = {}
            ttk.Label(win2, text="Adresa").grid(row=0, column=0, padx=5, pady=3)
            ttk.Label(win2, text="Discount (%)").grid(row=0, column=1, padx=5, pady=3)
            for i, row in df.iterrows():
                ttk.Label(win2, text=row["address"]).grid(
                    row=i + 1, column=0, padx=5, pady=2
                )
                e = ttk.Entry(win2, width=5)
                e.insert(0, "0")
                e.grid(row=i + 1, column=1, padx=5, pady=2)
                entry_map[row["id"]] = e

            def on_ok():
                # Aplic discount per locație (în fracțiune)
                discounts = {
                    loc_id: float(ent.get()) for loc_id, ent in entry_map.items()
                }
                df["% Discount"] = df["id"].map(discounts) / 100
                df["Discount Amount"] = df["Base Price"] * df["% Discount"]
                df["Final Price"] = df["Base Price"] - df["Discount Amount"]

                df_export = df[
                    [
                        "city",
                        "county",
                        "address",
                        "code",
                        "gps",
                        "photo_link",
                        "sqm",
                        "type",
                        "Base Price",
                        "% Discount",
                        "Final Price",
                        "Installation & Removal",
                        "Production",
                        "Availability",
                    ]
                ].copy()
                df_export.columns = [
                    "City",
                    "County",
                    "Address",
                    "CODE",
                    "GPS",
                    "Photo Link",
                    "SQM",
                    "Type",
                    "Base Price",
                    "% Discount",
                    "Final Price",
                    "Installation & Removal",
                    "Production",
                    "Availability",
                ]
                write_excel(df_export)
                win2.destroy()

            ttk.Button(win2, text="OK", command=on_ok).grid(
                row=len(df) + 1, column=0, columnspan=2, pady=10
            )

        else:
            # Discount global
            # Discount global (în fracțiune, pentru format % corect)
            df["% Discount"] = disc_pct / 100
            df["Discount Amount"] = df["Base Price"] * df["% Discount"]
            df["Final Price"] = df["Base Price"] - df["Discount Amount"]

            df_export = df[
                [
                    "city",
                    "county",
                    "address",
                    "code",
                    "gps",
                    "photo_link",
                    "sqm",
                    "type",
                    "Base Price",
                    "% Discount",
                    "Final Price",
                    "Installation & Removal",
                    "Production",
                    "Availability",
                ]
            ].copy()
            df_export.columns = [
                "City",
                "County",
                "Address",
                "CODE",
                "GPS",
                "Photo Link",
                "SQM",
                "Type",
                "Base Price",
                "% Discount",
                "Final Price",
                "Installation & Removal",
                "Production",
                "Availability",
            ]
            write_excel(df_export)

    # 9. Butonul de export (am mutat rândul la 5 pentru UI)
    ttk.Button(win, text="Generează Excel", command=export).grid(
        row=5, column=0, columnspan=2, pady=10
    )


def open_add_client_window(parent, refresh_cb=None):
    """Dialog pentru adăugarea unui client în tabela ``clienti``."""
    win = tk.Toplevel(parent)
    win.title("Adaugă client")

    # Nu mai solicităm detalii despre persoana de contact aici deoarece
    # imediat după crearea clientului este deschis automat dialogul pentru
    # adăugarea unei persoane de contact. Astfel evităm completarea
    # repetată a aceloraşi informaţii.
    labels = [
        "Nume companie",
        "CUI",
        "Adresă facturare",
        "Observații",
    ]
    entries = {}
    for i, lbl in enumerate(labels):
        ttk.Label(win, text=lbl + ":").grid(row=i, column=0, sticky="e", padx=5, pady=2)
        e = ttk.Entry(win, width=40)
        e.grid(row=i, column=1, padx=5, pady=2)
        entries[lbl] = e

    ttk.Label(win, text="Tip:").grid(
        row=len(labels), column=0, sticky="e", padx=5, pady=2
    )
    cb_tip = ttk.Combobox(win, values=["direct", "agency"], state="readonly", width=37)
    cb_tip.current(0)
    cb_tip.grid(row=len(labels), column=1, padx=5, pady=2)

    def save():
        nume = entries["Nume companie"].get().strip()
        if not nume:
            messagebox.showwarning("Lipsește numele", "Completează numele companiei.")
            return
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO clienti (nume, cui, adresa, observatii, tip) VALUES (?,?,?,?,?)",
            (
                nume,
                entries["CUI"].get().strip(),
                entries["Adresă facturare"].get().strip(),
                entries["Observații"].get().strip(),
                cb_tip.get() or "direct",
            ),
        )
        cid = cur.lastrowid
        if cid is None:
            cid = cur.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit()
        if refresh_cb:
            refresh_cb()
        win.destroy()
        parent.after(0, lambda: open_add_contact_window(parent, cid, refresh_cb))

    ttk.Button(win, text="Salvează", command=save).grid(
        row=len(labels) + 1, column=0, columnspan=2, pady=10
    )


def open_edit_client_window(parent, client_id, refresh_cb=None):
    cur = conn.cursor()
    row = cur.execute(
        "SELECT nume, cui, adresa, observatii, tip FROM clienti WHERE id=?",
        (client_id,),
    ).fetchone()
    if not row:
        return

    win = tk.Toplevel(parent)
    win.title("Editează client")

    labels = [
        "Nume companie",
        "CUI",
        "Adresă facturare",
        "Observații",
    ]
    entries = {}
    for i, lbl in enumerate(labels):
        ttk.Label(win, text=lbl + ":").grid(row=i, column=0, sticky="e", padx=5, pady=2)
        e = ttk.Entry(win, width=40)
        e.grid(row=i, column=1, padx=5, pady=2)
        entries[lbl] = e

    vals = list(row)
    for lbl, val in zip(labels, vals):
        entries[lbl].insert(0, val or "")

    ttk.Label(win, text="Tip:").grid(row=len(labels), column=0, sticky="e", padx=5, pady=2)
    cb_tip = ttk.Combobox(win, values=["direct", "agency"], state="readonly", width=37)
    cb_tip.grid(row=len(labels), column=1, padx=5, pady=2)
    cb_tip.set(row[4] or "direct")

    def save():
        nume = entries["Nume companie"].get().strip()
        if not nume:
            messagebox.showwarning("Lipsește numele", "Completează numele companiei.")
            return
        cur.execute(
            """
            UPDATE clienti SET nume=?, cui=?, adresa=?, observatii=?, tip=?
            WHERE id=?
            """,
            (
                nume,
                entries["CUI"].get().strip(),
                entries["Adresă facturare"].get().strip(),
                entries["Observații"].get().strip(),
                cb_tip.get() or "direct",
                client_id,
            ),
        )
        conn.commit()
        if refresh_cb:
            refresh_cb()
        win.destroy()

    ttk.Button(win, text="Salvează", command=save).grid(row=len(labels) + 1, column=0, columnspan=2, pady=10)


def open_add_contact_window(parent, client_id, refresh_cb=None):
    win = tk.Toplevel(parent)
    win.title("Adaugă persoană de contact")

    labels = ["Nume", "Rol", "Email", "Telefon"]
    entries = {}
    for i, lbl in enumerate(labels):
        ttk.Label(win, text=lbl + ":").grid(row=i, column=0, sticky="e", padx=5, pady=2)
        e = ttk.Entry(win, width=40)
        e.grid(row=i, column=1, padx=5, pady=2)
        entries[lbl] = e

    def save():
        name = entries["Nume"].get().strip()
        if not name:
            messagebox.showwarning("Lipsește numele", "Completează numele.")
            return
        add_client_contact(
            client_id,
            name,
            entries["Rol"].get().strip(),
            entries["Email"].get().strip(),
            entries["Telefon"].get().strip(),
        )
        if refresh_cb:
            refresh_cb()
        win.destroy()

    ttk.Button(win, text="Salvează", command=save).grid(row=len(labels), column=0, columnspan=2, pady=10)


def open_edit_contact_window(parent, contact: dict, refresh_cb=None):
    """Edit an existing contact."""
    win = tk.Toplevel(parent)
    win.title("Editează persoană de contact")

    labels = ["Nume", "Rol", "Email", "Telefon"]
    entries = {}
    values = [contact.get("nume", ""), contact.get("rol", ""), contact.get("email", ""), contact.get("phone", "")]
    for i, (lbl, val) in enumerate(zip(labels, values)):
        ttk.Label(win, text=lbl + ":").grid(row=i, column=0, sticky="e", padx=5, pady=2)
        e = ttk.Entry(win, width=40)
        e.grid(row=i, column=1, padx=5, pady=2)
        e.insert(0, val)
        entries[lbl] = e

    def save():
        name = entries["Nume"].get().strip()
        if not name:
            messagebox.showwarning("Lipsește numele", "Completează numele.")
            return
        update_client_contact(
            contact["id"],
            name,
            entries["Rol"].get().strip(),
            entries["Email"].get().strip(),
            entries["Telefon"].get().strip(),
        )
        if refresh_cb:
            refresh_cb()
        win.destroy()

    ttk.Button(win, text="Salvează", command=save).grid(row=len(labels), column=0, columnspan=2, pady=10)


def open_contact_window(parent, contact: dict, refresh_cb=None) -> None:
    """Display contact details in a dedicated window."""
    win = tk.Toplevel(parent)
    win.title(contact.get("nume", "Contact"))
    default_font = tkfont.nametofont("TkDefaultFont")
    big_font = default_font.copy()
    big_font.configure(size=14)

    fields = [
        ("Nume", contact.get("nume", "")),
        ("Rol", contact.get("rol", "")),
        ("Email", contact.get("email", "")),
        ("Telefon", contact.get("phone", "")),
    ]

    for i, (lbl, val) in enumerate(fields):
        ttk.Label(win, text=lbl + ":", font=(default_font.cget("family"), 12, "bold")).grid(
            row=i, column=0, sticky="e", padx=5, pady=5
        )
        ttk.Label(win, text=val, font=big_font).grid(row=i, column=1, sticky="w", padx=5, pady=5)
        if lbl in ("Email", "Telefon") and val:
            def _copy(text=val):
                win.clipboard_clear()
                win.clipboard_append(text)

            ttk.Button(win, text="Copiază", command=_copy).grid(row=i, column=2, padx=5, pady=5)

    btn_frame = ttk.Frame(win)
    btn_frame.grid(row=len(fields), column=0, columnspan=3, pady=10)

    def edit_contact():
        win.destroy()
        open_edit_contact_window(parent, contact, refresh_cb)

    def delete_contact():
        if messagebox.askyesno("Confirmă", "Ștergi acest contact?"):
            delete_client_contact(contact["id"])
            if refresh_cb:
                refresh_cb()
            win.destroy()

    ttk.Button(btn_frame, text="Editează", command=edit_contact).grid(row=0, column=0, padx=5)
    ttk.Button(btn_frame, text="Șterge", command=delete_contact).grid(row=0, column=1, padx=5)



def open_client_detail(tree, event):
    rowid = tree.identify_row(event.y)
    if not rowid:
        return
    cid = int(rowid)
    cur = conn.cursor()
    row = cur.execute(
        "SELECT nume, tip, cui, adresa, observatii FROM clienti WHERE id=?",
        (cid,),
    ).fetchone()
    if not row:
        return

    win = tk.Toplevel(tree.master)
    win.title(f"Client {row[0]}")

    labels = [
        ("Nume", row[0]),
        ("Tip", row[1] or "direct"),
        ("CUI", row[2] or ""),
        ("Adresă", row[3] or ""),
        ("Observații", row[4] or ""),
    ]

    for i, (lbl, val) in enumerate(labels):
        ttk.Label(win, text=lbl + ":", font=("Segoe UI", 9, "bold")).grid(row=i, column=0, sticky="e", padx=5, pady=2)
        ttk.Label(win, text=val).grid(row=i, column=1, sticky="w", padx=5, pady=2)

    contacts = get_client_contacts(cid)
    start_row = len(labels) + 1
    if contacts:
        ttk.Label(win, text="Persoane contact:").grid(row=start_row, column=0, sticky="ne", padx=5, pady=5)
        frm = ttk.Frame(win)
        frm.grid(row=start_row, column=1, sticky="w")
        for j, c in enumerate(contacts):
            text = c["nume"] if not c.get("rol") else f"{c['nume']} ({c['rol']})"
            lbl = ttk.Label(frm, text=text, foreground="blue", cursor="hand2")
            lbl.grid(row=j, column=0, sticky="w")
            lbl.bind(
                "<Button-1>",
                lambda e, c=c: open_contact_window(
                    win, c, lambda: (win.destroy(), open_client_detail(tree, event))
                ),
            )

    def add_contact():
        open_add_contact_window(win, cid, lambda: (win.destroy(), open_client_detail(tree, event)))

    ttk.Button(win, text="Adaugă persoană", command=add_contact).grid(row=start_row + len(contacts) + 1, column=0, columnspan=2, pady=10)



def _write_backup_excel(rows, start_m: datetime.date, end_m: datetime.date, path: str) -> None:
    """Write an Excel backup file using ``openpyxl``."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    if not rows:
        return

    days_in_month = (end_m - start_m).days + 1

    data_rows = []
    header_info = None
    header_start = None
    header_end = None
    idx = 1
    for (
        client_name,
        client_cui,
        client_addr,
        firma_name,
        firma_cui,
        firma_addr,
        campaign,
        city,
        addr,
        code,
        face,
        typ,
        size,
        sqm,
        ds,
        de,
        price,
        deco_cost_loc,
        cid,
        deco_r,
        prod_r,
    ) in rows:
        ds_dt = datetime.date.fromisoformat(ds)
        de_dt = datetime.date.fromisoformat(de)
        if header_start is None or ds_dt < header_start:
            header_start = ds_dt
        if header_end is None or de_dt > header_end:
            header_end = de_dt
        if header_info is None:
            header_info = (
                firma_name,
                firma_cui,
                firma_addr,
                client_name,
                client_cui,
                client_addr,
                campaign,
            )
        ov_start = max(ds_dt, start_m)
        ov_end = min(de_dt, end_m)
        days = (ov_end - ov_start).days + 1
        frac = days / days_in_month
        amount = price * frac
        deco = deco_r if deco_r is not None else (deco_cost_loc or 0.0)
        try:
            prod_default = round(float(sqm or 0) * 7, 2)
        except Exception:
            prod_default = 0.0
        prod = prod_r if prod_r is not None else prod_default
        data_rows.append([
            idx,
            city,
            addr,
            code,
            face,
            1,
            typ,
            size,
            ov_start,
            ov_end,
            round(frac, 2),
            "EUR",
            price,
            amount,
            deco,
            prod,
        ])
        idx += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Backup"

    dark_blue = PatternFill(fill_type="solid", fgColor="305496")
    white_bold = Font(color="FFFFFF", bold=True)

    if header_info:
        f_name, f_cui, f_addr, c_name, c_cui, c_addr, camp = header_info
    else:
        f_name = f_cui = f_addr = c_name = c_cui = c_addr = camp = ""

    ws.merge_cells("A1:P1")
    title_cell = ws["A1"]
    title_cell.value = f"BKP {f_name} x {c_name} - {camp or c_name} - {start_m:%B}"
    title_cell.fill = dark_blue
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, 6):
        for c in range(1, 17):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A2:E2"); ws["A2"].value = f"Societatea care facturează: {f_name}"
    ws.merge_cells("A3:E3"); ws["A3"].value = f"CUI: {f_cui}"
    ws.merge_cells("A4:E4"); ws["A4"].value = f"Adresă: {f_addr}"

    ws.merge_cells("F2:J2"); ws["F2"].value = f"Client: {c_name}"
    ws.merge_cells("F3:J3"); ws["F3"].value = f"CUI client: {c_cui}"
    ws.merge_cells("F4:J4"); ws["F4"].value = f"Adresă client: {c_addr}"

    if header_start is None:
        header_start = start_m
    if header_end is None:
        header_end = end_m
    ws.merge_cells("K2:P2"); ws["K2"].value = f"Perioada campanie: {header_start:%d.%m.%Y} - {header_end:%d.%m.%Y}"
    ws.merge_cells("K3:P3"); ws["K3"].value = f"Denumire campanie: {camp or c_name}"

    headers = [
        "Nr. Crt",
        "Oraș",
        "Adresă",
        "Cod",
        "Cod Față",
        "Număr bucăți",
        "Tip Suport",
        "Dimensiune",
        "Data Început",
        "Data Sfârșit",
        "Perioada (luni)",
        "Valută",
        "Preț chirie/lună",
        "Chirie NET",
        "Preț Decorare",
        "Preț Producție",
    ]

    header_row = 5
    for idx_col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx_col, value=header)
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_start = header_row + 1
    currency_cols = {13, 14, 15, 16}
    for r_off, row_vals in enumerate(data_rows):
        r = data_start + r_off
        for c_off, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c_off, value=val)
            if isinstance(val, datetime.date):
                cell.number_format = "DD.MM.YYYY"
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif isinstance(val, (int, float)):
                if c_off == 1:
                    cell.number_format = "0"
                elif c_off in currency_cols:
                    cell.number_format = "\u20ac#,##0.00"
                else:
                    cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    last_data_row = data_start + len(data_rows) - 1
    total_start = last_data_row + 1

    # total lines for each cost column
    deco_total = ws.cell(row=total_start, column=15, value=f"=SUM(O{data_start}:O{last_data_row})")
    deco_total.font = Font(bold=True)
    deco_total.number_format = "\u20ac#,##0.00"
    deco_total.alignment = Alignment(horizontal="center", vertical="center")

    prod_total = ws.cell(row=total_start, column=16, value=f"=SUM(P{data_start}:P{last_data_row})")
    prod_total.font = Font(bold=True)
    prod_total.number_format = "\u20ac#,##0.00"
    prod_total.alignment = Alignment(horizontal="center", vertical="center")

    chirie_total = ws.cell(row=total_start, column=14, value=f"=SUM(N{data_start}:N{last_data_row})")
    chirie_total.font = Font(bold=True)
    chirie_total.number_format = "\u20ac#,##0.00"
    chirie_total.alignment = Alignment(horizontal="center", vertical="center")

    grand_row = total_start + 1
    # move the total label one column to the right so it sits next to the
    # computed sum value
    lbl = ws.cell(row=grand_row, column=15, value="Total")
    # do not merge the label cell so it occupies a single box
    lbl.font = Font(bold=True)
    lbl.alignment = Alignment(horizontal="center", vertical="center")
    val = ws.cell(
        row=grand_row,
        column=16,
        value=f"=SUM(N{data_start}:N{last_data_row})+SUM(O{data_start}:O{last_data_row})+SUM(P{data_start}:P{last_data_row})",
    )
    val.font = Font(bold=True)
    val.number_format = "\u20ac#,##0.00"
    val.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(border_style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    last_total_row = grand_row
    for r in range(1, last_total_row + 1):
        for c in range(1, 17):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None or cell.coordinate in ws.merged_cells:
                cell.border = border

    # apply a thicker border only to the "Total" label and final sum cells
    thick = Side(border_style="thick")
    start_col, end_col = 15, 16
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=grand_row, column=c)
        left = thick if c == start_col else thin
        right = thick if c == end_col else thin
        cell.border = Border(top=thick, bottom=thick, left=left, right=right)

    from openpyxl.utils import get_column_letter
    extra_height_needed = False
    for col_idx in range(1, 17):
        max_len = 0
        start_row = 2 if col_idx == 1 else 1
        for row in range(start_row, last_total_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max_len + 2, 25)
        if width > 20:
            width = 20
            extra_height_needed = True
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    data_height = 35 if extra_height_needed else 30
    for r in range(data_start, last_data_row + 1):
        ws.row_dimensions[r].height = data_height
    for r in range(last_data_row + 1, last_total_row + 1):
        ws.row_dimensions[r].height = 22 if not extra_height_needed else 26

    for r in range(1, header_row + 1):
        ws.row_dimensions[r].height = 32 if extra_height_needed else 30

    wb.save(path)
def export_client_backup(month, year, client_id=None, firma_id=None, campaign=None, directory=None):
    """Exportă un backup de facturare pentru luna dată, formatat în Excel."""
    import calendar
    import os

    days_in_month = calendar.monthrange(year, month)[1]
    start_m = datetime.date(year, month, 1)
    end_m = datetime.date(year, month, days_in_month)

    cur = conn.cursor()
    sql = (
        "SELECT c.nume, c.cui, c.adresa, "
        "       f.nume, f.cui, f.adresa, r.campaign, "
        "       l.city, l.address, l.code, l.face, l.type, l.size, l.sqm, "
        "       r.data_start, r.data_end, r.suma, l.decoration_cost, r.client_id, "
        "       r.decor_cost, r.prod_cost, r.id, r.loc_id "
        "FROM rezervari r "
        "JOIN locatii l ON r.loc_id = l.id "
        "JOIN clienti c ON r.client_id = c.id "
        "LEFT JOIN firme f ON r.firma_id = f.id "
        "WHERE r.suma IS NOT NULL AND r.suma > 0"
        " AND NOT (r.data_end < ? OR r.data_start > ?)"
    )
    params = [start_m.isoformat(), end_m.isoformat()]
    if client_id:
        sql += " AND r.client_id=?"
        params.append(client_id)
    if firma_id:
        sql += " AND r.firma_id=?"
        params.append(firma_id)
    if campaign:
        sql += " AND IFNULL(r.campaign, '')=?"
        params.append(campaign or "")

    rows = cur.execute(sql, params).fetchall()
    if not rows:
        messagebox.showinfo("Export", "Nu există închirieri pentru perioada aleasă.")
        return

    processed = []
    for row in rows:
        (
            client_name,
            client_cui,
            client_addr,
            firma_name,
            firma_cui,
            firma_addr,
            campaign,
            city,
            addr,
            code,
            face,
            typ,
            size,
            sqm,
            ds,
            de,
            price,
            deco_cost_loc,
            cid,
            deco_r,
            prod_r,
            rez_id,
            loc_id,
        ) = row

        ds_dt = datetime.date.fromisoformat(ds)
        de_dt = datetime.date.fromisoformat(de)
        deco = prod = 0.0
        if start_m <= ds_dt <= end_m:
            deco = deco_r if deco_r is not None else (deco_cost_loc or 0.0)
            try:
                prod_default = round(float(sqm or 0) * 7, 2)
            except Exception:
                prod_default = 0.0
            prod = prod_r if prod_r is not None else prod_default

        cur.execute(
            """
            SELECT COALESCE(SUM(decor_cost),0), COALESCE(SUM(prod_cost),0)
              FROM decorari
             WHERE loc_id=? AND data BETWEEN ? AND ?
               AND data BETWEEN ? AND ?
               AND (rez_id=? OR rez_id IS NULL)
            """,
            (
                loc_id,
                start_m.isoformat(),
                end_m.isoformat(),
                ds,
                de,
                rez_id,
            ),
        )
        extra_deco, extra_prod = cur.fetchone() or (0.0, 0.0)
        deco += extra_deco or 0.0
        prod += extra_prod or 0.0

        processed.append(
            (
                client_name,
                client_cui,
                client_addr,
                firma_name,
                firma_cui,
                firma_addr,
                campaign,
                city,
                addr,
                code,
                face,
                typ,
                size,
                sqm,
                ds,
                de,
                price,
                deco_cost_loc,
                cid,
                deco,
                prod,
            )
        )

    if directory is None:
        directory = filedialog.askdirectory()
        if not directory:
            return

    month_dir = os.path.join(directory, f"BKP {start_m:%B %Y}")
    os.makedirs(month_dir, exist_ok=True)

    groups = {}
    for row in processed:
        f_name = row[3] or "FaraFirma"
        c_name = row[0] or ""
        camp = row[6] or c_name
        key = (f_name, c_name, camp)
        groups.setdefault(key, []).append(row)

    for (f_name, c_name, camp), grp_rows in groups.items():
        sub = os.path.join(month_dir, _safe_filename(f_name))
        os.makedirs(sub, exist_ok=True)
        fname = _safe_filename(f"BKP {f_name} x {c_name} - {camp} - {start_m:%B}.xlsx")
        path = os.path.join(sub, fname)
        _write_backup_excel(grp_rows, start_m, end_m, path)

    messagebox.showinfo("Export", f"Backupurile au fost salvate în:\n{month_dir}")


def export_all_backups(month, year):
    """Generează backupuri pentru toți clienții cu închiriere activă."""
    import calendar
    import os

    days_in_month = calendar.monthrange(year, month)[1]
    start_m = datetime.date(year, month, 1)
    end_m = datetime.date(year, month, days_in_month)

    cur = conn.cursor()
    sql = (
        "SELECT c.nume, c.cui, c.adresa, "
        "       f.nume, f.cui, f.adresa, r.campaign, "
        "       l.city, l.address, l.code, l.face, l.type, l.size, l.sqm, "
        "       r.data_start, r.data_end, r.suma, l.decoration_cost, r.client_id, "
        "       r.decor_cost, r.prod_cost, r.id, r.loc_id "
        "FROM rezervari r "
        "JOIN locatii l ON r.loc_id = l.id "
        "JOIN clienti c ON r.client_id = c.id "
        "LEFT JOIN firme f ON r.firma_id = f.id "
        "WHERE r.suma IS NOT NULL AND r.suma > 0"
        " AND NOT (r.data_end < ? OR r.data_start > ?)"
    )
    rows = cur.execute(sql, (start_m.isoformat(), end_m.isoformat())).fetchall()
    if not rows:
        messagebox.showinfo("Export", "Nu există închirieri pentru perioada aleasă.")
        return

    processed = []
    for row in rows:
        (
            client_name,
            client_cui,
            client_addr,
            firma_name,
            firma_cui,
            firma_addr,
            campaign,
            city,
            addr,
            code,
            face,
            typ,
            size,
            sqm,
            ds,
            de,
            price,
            deco_cost_loc,
            cid,
            deco_r,
            prod_r,
            rez_id,
            loc_id,
        ) = row

        ds_dt = datetime.date.fromisoformat(ds)
        de_dt = datetime.date.fromisoformat(de)
        deco = prod = 0.0
        if start_m <= ds_dt <= end_m:
            deco = deco_r if deco_r is not None else (deco_cost_loc or 0.0)
            try:
                prod_default = round(float(sqm or 0) * 7, 2)
            except Exception:
                prod_default = 0.0
            prod = prod_r if prod_r is not None else prod_default

        cur.execute(
            """
            SELECT COALESCE(SUM(decor_cost),0), COALESCE(SUM(prod_cost),0)
              FROM decorari
             WHERE loc_id=? AND data BETWEEN ? AND ?
               AND data BETWEEN ? AND ?
               AND (rez_id=? OR rez_id IS NULL)
            """,
            (
                loc_id,
                start_m.isoformat(),
                end_m.isoformat(),
                ds,
                de,
                rez_id,
            ),
        )
        extra_deco, extra_prod = cur.fetchone() or (0.0, 0.0)
        deco += extra_deco or 0.0
        prod += extra_prod or 0.0

        processed.append(
            (
                client_name,
                client_cui,
                client_addr,
                firma_name,
                firma_cui,
                firma_addr,
                campaign,
                city,
                addr,
                code,
                face,
                typ,
                size,
                sqm,
                ds,
                de,
                price,
                deco_cost_loc,
                cid,
                deco,
                prod,
            )
        )

    base_dir = filedialog.askdirectory()
    if not base_dir:
        return

    month_dir = os.path.join(base_dir, f"BKP {start_m:%B %Y}")
    os.makedirs(month_dir, exist_ok=True)

    groups = {}
    for row in processed:
        f_name = row[3] or "FaraFirma"
        c_name = row[0] or ""
        camp = row[6] or c_name
        key = (f_name, c_name, camp)
        groups.setdefault(key, []).append(row)

    for (f_name, c_name, camp), grp_rows in groups.items():
        sub = os.path.join(month_dir, _safe_filename(f_name))
        os.makedirs(sub, exist_ok=True)
        fname = _safe_filename(f"BKP {f_name} x {c_name} - {camp} - {start_m:%B}.xlsx")
        path = os.path.join(sub, fname)
        _write_backup_excel(grp_rows, start_m, end_m, path)

    messagebox.showinfo("Export", f"Backupurile au fost salvate în:\n{month_dir}")


def open_clients_window(root, user=None):
    """Fereastră pentru administrarea clienților.

    ``user`` is optional and defaults to an admin-like role for backwards
    compatibility.
    """
    win = tk.Toplevel(root)
    win.title("Clienți")

    ttk.Label(win, text="Caută:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    search_var = tk.StringVar()
    entry_search = ttk.Entry(win, textvariable=search_var, width=20)
    entry_search.grid(row=0, column=1, columnspan=3, sticky="ew", padx=5, pady=5)

    cols = ("Nume", "Tip", "Contact", "Email", "Telefon", "CUI", "Adresă", "Observații")
    tree = ttk.Treeview(win, columns=cols, show="headings")
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=120, anchor="w")
    tree.grid(row=1, column=0, columnspan=4, sticky="nsew")

    vsb = ttk.Scrollbar(win, orient="vertical", command=tree.yview)
    vsb.grid(row=1, column=4, sticky="ns")
    tree.configure(yscroll=vsb.set)

    win.columnconfigure(1, weight=1)
    win.rowconfigure(1, weight=1)

    def refresh():
        tree.delete(*tree.get_children())
        term = search_var.get().strip().lower()
        rows = (
            conn.cursor()
            .execute(
                """
                SELECT c.id, c.nume, c.tip,
                       COALESCE(cc.nume, c.contact) AS contact,
                       COALESCE(cc.email, c.email) AS email,
                       COALESCE(cc.phone, c.phone) AS phone,
                       c.cui, c.adresa, c.observatii
                  FROM clienti c
             LEFT JOIN client_contacts cc
                    ON cc.client_id = c.id AND cc.id = (
                        SELECT id FROM client_contacts c2
                         WHERE c2.client_id = c.id
                         ORDER BY c2.id LIMIT 1
                    )
              ORDER BY c.nume
                """
            )
            .fetchall()
        )
        for cid, nume, tip, contact, email, phone, cui, addr, obs in rows:
            if term and not (
                term in (nume or "").lower()
                or term in (contact or "").lower()
                or term in (email or "").lower()
            ):
                continue
            tree.insert(
                "",
                "end",
                iid=str(cid),
                values=(nume, tip or "direct", contact, email, phone, cui or "", addr or "", obs or ""),
            )

    def add_client():
        open_add_client_window(win, refresh)

    def delete_client():
        sel = tree.selection()
        if not sel:
            return
        if not messagebox.askyesno("Confirmă", "Ștergi clientul selectat?"):
            return
        cid = int(sel[0])
        cur = conn.cursor()
        nume_row = cur.execute(
            "SELECT nume FROM clienti WHERE id=?",
            (cid,),
        ).fetchone()
        if not nume_row:
            return
        nume_cli = nume_row[0]

        rentals = cur.execute(
            "SELECT COUNT(*) FROM rezervari WHERE client_id=? OR client=?",
            (cid, nume_cli),
        ).fetchone()[0]
        if rentals:
            messagebox.showerror(
                "Ștergere interzisă",
                "Clientul are rezervări înregistrate și nu poate fi șters.",
            )
            return

        cur.execute("DELETE FROM client_contacts WHERE client_id=?", (cid,))
        cur.execute("DELETE FROM clienti WHERE id=?", (cid,))
        conn.commit()
        refresh()

    def export_current():
        sel = tree.selection()
        cid = int(sel[0]) if sel else None
        today = datetime.date.today()
        year = simpledialog.askinteger(
            "An", "Anul", initialvalue=today.year, parent=win
        )
        if year is None:
            return
        month = simpledialog.askinteger(
            "Luna",
            "Luna (1-12)",
            minvalue=1,
            maxvalue=12,
            initialvalue=today.month,
            parent=win,
        )
        if month is None:
            return
        export_client_backup(month, year, cid)

    def export_all():
        today = datetime.date.today()
        year = simpledialog.askinteger("An", "Anul", initialvalue=today.year, parent=win)
        if year is None:
            return
        month = simpledialog.askinteger(
            "Luna",
            "Luna (1-12)",
            minvalue=1,
            maxvalue=12,
            initialvalue=today.month,
            parent=win,
        )
        if month is None:
            return
        export_all_backups(month, year)

    btn_add = ttk.Button(win, text="Adaugă", command=add_client)
    btn_edit = ttk.Button(win, text="Editează", command=lambda: open_edit_client_window(win, int(tree.selection()[0]), refresh))
    btn_del = ttk.Button(win, text="Șterge", command=delete_client)
    btn_export = ttk.Button(win, text="Export Backup", command=export_current)
    btn_export_all = ttk.Button(win, text="Export Toți", command=export_all)
    for i, b in enumerate((btn_add, btn_edit, btn_del, btn_export, btn_export_all)):
        b.grid(row=2, column=i, padx=5, pady=5, sticky="w")

    def on_sel(event=None):
        sel = tree.selection()
        state = "normal" if sel else "disabled"
        btn_edit.config(state=state)
        btn_del.config(state=state)

    tree.bind("<<TreeviewSelect>>", on_sel)
    on_sel()

    tree.bind("<Double-1>", lambda e: open_client_detail(tree, e))

    def on_search_change(*args):
        refresh()

    search_var.trace_add("write", on_search_change)

    refresh()


def export_vendor_report(month=None, year=None):
    """Exportă un raport detaliat pentru fiecare vânzător.

    Parametrii ``month`` și ``year`` indică luna pentru care se generează
    raportul. Dacă nu sunt precizați, se folosește luna și anul curent.
    """
    import pandas as pd
    from tkinter import filedialog, messagebox
    from db import read_sql_query

    today = datetime.date.today()
    if year is None:
        year = today.year
    if month is None:
        month = today.month

    report_date = datetime.date(year, month, 1)
    current_month = report_date.strftime("%B")
    special_june = (month == 6 and year == 2025)

    users = read_sql_query(
        "SELECT username, comune FROM users WHERE role='seller'",
    )
    if users.empty:
        messagebox.showinfo("Raport", "Nu există vânzători.")
        return

    df = read_sql_query(
        """
        SELECT r.created_by, r.created_on, r.client, r.suma, r.data_start, r.data_end,
               l.city, l.county, l.address
          FROM rezervari r
          JOIN locatii l ON r.loc_id = l.id
         WHERE r.suma IS NOT NULL AND r.suma > 0
        """,
        parse_dates=["created_on", "data_start", "data_end"],
    )

    if df.empty:
        messagebox.showinfo("Raport", "Nu există închirieri.")
        return

    import calendar
    month_end = datetime.date(year, month, calendar.monthrange(year, month)[1])
    df = df[
        (df["created_on"].dt.date >= report_date)
        & (df["created_on"].dt.date <= month_end)
        & (df["data_start"] >= df["created_on"])
    ]

    if df.empty:
        messagebox.showinfo(
            "Raport", "Nu există închirieri active în perioada aleasă.")
        return

    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        title="Salvează raportul",
    )
    if not path:
        return

    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        wb = writer.book
        hdr_fmt = wb.add_format(
            {
                "bold": True,
                "bg_color": "#4F81BD",
                "font_color": "white",
                "align": "center",
            }
        )
        money_fmt = wb.add_format({"num_format": "€#,##0.00", "align": "center"})
        center_fmt = wb.add_format({"align": "center"})

        if special_june:
            month_end_date = datetime.date(year, month, calendar.monthrange(year, month)[1])

        def contract_months(ds: datetime.date, de: datetime.date) -> float:
            cur = ds
            total = 0.0
            while cur <= de:
                dim = calendar.monthrange(cur.year, cur.month)[1]
                month_end = datetime.date(cur.year, cur.month, dim)
                ov_end = min(de, month_end)
                total += ((ov_end - cur).days + 1) / dim
                cur = ov_end + datetime.timedelta(days=1)
            return total

        def split_by_month(
            ds: datetime.date, de: datetime.date, price_per_month: float
        ):
            cur = ds
            while cur <= de:
                dim = calendar.monthrange(cur.year, cur.month)[1]
                month_end = datetime.date(cur.year, cur.month, dim)
                ov_end = min(de, month_end)
                frac = ((ov_end - cur).days + 1) / dim
                yield cur.strftime("%B"), price_per_month * frac
                cur = ov_end + datetime.timedelta(days=1)

        for _, row in users.iterrows():
            uname = row["username"]
            comune = {c for c in (row["comune"] or "").split(",") if c}
            sub = df[df["created_by"] == uname]
            if comune:
                sub = sub[sub["county"].isin(comune)]
            if sub.empty:
                continue

            sub = sub.copy()
            sub["Luni"] = sub.apply(
                lambda r: contract_months(r["data_start"].date(), r["data_end"].date()),
                axis=1,
            )
            sub["Chirie/lună"] = sub["suma"]
            sub["Valoare"] = sub["suma"] * sub["Luni"]
            sub["Perioadă"] = sub.apply(
                lambda r: f"{r['data_start'].date()} → {r['data_end'].date()}",
                axis=1,
            )

            df_det = sub[
                [
                    "city",
                    "county",
                    "address",
                    "client",
                    "Perioadă",
                    "Luni",
                    "Chirie/lună",
                    "Valoare",
                ]
            ].copy()
            df_det.columns = [
                "Oraș",
                "Județ",
                "Adresă",
                "Client",
                "Perioadă",
                "Luni",
                "Chirie/lună",
                "Valoare",
            ]
            df_det.insert(0, "Nr.crt", range(1, len(df_det) + 1))

            sheet = uname[:31]
            df_det.to_excel(writer, sheet_name=sheet, index=False, startrow=0)

            ws = writer.sheets[sheet]
            for col_idx, col in enumerate(df_det.columns):
                width = max(len(str(col)), df_det[col].astype(str).map(len).max()) + 2
                fmt = money_fmt if col in {"Chirie/lună", "Valoare"} else center_fmt
                ws.set_column(col_idx, col_idx, width, fmt)
                ws.write(0, col_idx, col, hdr_fmt)

            total_row = len(df_det) + 1
            ws.write(total_row, df_det.columns.get_loc("Client"), "Total", hdr_fmt)
            ws.write(
                total_row,
                df_det.columns.get_loc("Valoare"),
                df_det["Valoare"].sum(),
                money_fmt,
            )

    messagebox.showinfo("Raport", f"Raport salvat:\n{path}")


def open_users_window(root):
    """Admin window to manage user accounts."""

    win = tk.Toplevel(root)
    win.title("Utilizatori")

    tree = ttk.Treeview(win, columns=("User", "Role", "Comune"), show="headings")
    for c in ("User", "Role", "Comune"):
        tree.heading(c, text=c)
    tree.grid(row=0, column=0, columnspan=3, sticky="nsew")
    win.columnconfigure(0, weight=1)
    win.rowconfigure(0, weight=1)

    def refresh():
        tree.delete(*tree.get_children())
        rows = (
            conn.cursor().execute("SELECT username, role, comune FROM users").fetchall()
        )
        for u, r, c in rows:
            tree.insert("", "end", values=(u, r, c or ""))

    def add_user():
        dlg = tk.Toplevel(win)
        dlg.title("Adaugă utilizator")

        ttk.Label(dlg, text="Utilizator:").grid(
            row=0, column=0, padx=5, pady=5, sticky="e"
        )
        entry_user = ttk.Entry(dlg, width=30)
        entry_user.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(dlg, text="Parola:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        entry_pass = ttk.Entry(dlg, width=30, show="*")
        entry_pass.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(dlg, text="Rol:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        combo_role = ttk.Combobox(dlg, values=["admin", "manager", "seller"], state="readonly")
        combo_role.current(1)
        combo_role.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(dlg, text="Comune:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
        entry_comune = ttk.Entry(dlg, width=30)
        entry_comune.grid(row=3, column=1, padx=5, pady=5)

        def on_ok():
            u = entry_user.get().strip()
            if not u:
                messagebox.showwarning(
                    "Date lipsă", "Completează utilizatorul.", parent=dlg
                )
                return
            p = entry_pass.get()
            if p is None:
                messagebox.showwarning("Date lipsă", "Completează parola.", parent=dlg)
                return
            role = combo_role.get() or "seller"
            comune = entry_comune.get().strip() or ""
            create_user(u, p, role, comune)
            dlg.destroy()
            refresh()

        ttk.Button(dlg, text="Salvează", command=on_ok).grid(
            row=4, column=0, columnspan=2, pady=10
        )
        dlg.grab_set()
        entry_user.focus()

    def delete_user():
        sel = tree.selection()
        if not sel:
            return
        user = tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirmă", f"Ștergi utilizatorul {user}?"):
            conn.cursor().execute("DELETE FROM users WHERE username=?", (user,))
            conn.commit()
            refresh()

    ttk.Button(win, text="Adaugă", command=add_user).grid(row=1, column=0, pady=5)
    ttk.Button(win, text="Șterge", command=delete_user).grid(row=1, column=1, pady=5)
    ttk.Button(win, text="Închide", command=win.destroy).grid(row=1, column=2, pady=5)

    refresh()


def open_firme_window(root):
    """Admin window to manage invoice companies."""

    win = tk.Toplevel(root)
    win.title("Firme")

    tree = ttk.Treeview(win, columns=("Nume", "CUI", "Adresă"), show="headings")
    for c in ("Nume", "CUI", "Adresă"):
        tree.heading(c, text=c)
    tree.grid(row=0, column=0, columnspan=3, sticky="nsew")
    win.columnconfigure(0, weight=1)
    win.rowconfigure(0, weight=1)

    def refresh():
        tree.delete(*tree.get_children())
        rows = conn.cursor().execute("SELECT id, nume, cui, adresa FROM firme ORDER BY nume").fetchall()
        for fid, n, c, a in rows:
            tree.insert("", "end", iid=str(fid), values=(n, c or "", a or ""))

    def add_firma():
        dlg = tk.Toplevel(win)
        dlg.title("Adaugă firmă")

        lbls = ["Nume", "CUI", "Adresă"]
        ents = {}
        for i, l in enumerate(lbls):
            ttk.Label(dlg, text=l + ":").grid(row=i, column=0, padx=5, pady=5, sticky="e")
            e = ttk.Entry(dlg, width=40)
            e.grid(row=i, column=1, padx=5, pady=5)
            ents[l] = e

        def on_ok():
            name = ents["Nume"].get().strip()
            if not name:
                messagebox.showwarning("Date lipsă", "Completează numele.", parent=dlg)
                return
            conn.cursor().execute(
                "INSERT INTO firme (nume, cui, adresa) VALUES (?,?,?)",
                (name, ents["CUI"].get().strip(), ents["Adresă"].get().strip()),
            )
            conn.commit()
            dlg.destroy()
            refresh()

        ttk.Button(dlg, text="Salvează", command=on_ok).grid(row=len(lbls), column=0, columnspan=2, pady=10)
        dlg.grab_set()
        ents["Nume"].focus()

    def del_firma():
        sel = tree.selection()
        if not sel:
            return
        fid = int(sel[0])
        if messagebox.askyesno("Confirmă", "Ștergi firma selectată?"):
            conn.cursor().execute("DELETE FROM firme WHERE id=?", (fid,))
            conn.commit()
            refresh()

    def edit_firma():
        sel = tree.selection()
        if not sel:
            return
        fid = int(sel[0])
        row = conn.cursor().execute(
            "SELECT nume, cui, adresa FROM firme WHERE id=?", (fid,)
        ).fetchone()
        if not row:
            return
        dlg = tk.Toplevel(win)
        dlg.title("Editează firmă")

        lbls = ["Nume", "CUI", "Adresă"]
        ents = {}
        for i, l in enumerate(lbls):
            ttk.Label(dlg, text=l + ":").grid(row=i, column=0, padx=5, pady=5, sticky="e")
            e = ttk.Entry(dlg, width=40)
            e.grid(row=i, column=1, padx=5, pady=5)
            ents[l] = e

        ents["Nume"].insert(0, row[0])
        ents["CUI"].insert(0, row[1] or "")
        ents["Adresă"].insert(0, row[2] or "")

        def on_ok():
            name = ents["Nume"].get().strip()
            if not name:
                messagebox.showwarning("Date lipsă", "Completează numele.", parent=dlg)
                return
            conn.cursor().execute(
                "UPDATE firme SET nume=?, cui=?, adresa=? WHERE id=?",
                (name, ents["CUI"].get().strip(), ents["Adresă"].get().strip(), fid),
            )
            conn.commit()
            dlg.destroy()
            refresh()

        ttk.Button(dlg, text="Salvează", command=on_ok).grid(row=len(lbls), column=0, columnspan=2, pady=10)
        dlg.grab_set()
        ents["Nume"].focus()

    ttk.Button(win, text="Adaugă", command=add_firma).grid(row=1, column=0, pady=5)
    ttk.Button(win, text="Editează", command=edit_firma).grid(row=1, column=1, pady=5)
    ttk.Button(win, text="Șterge", command=del_firma).grid(row=1, column=2, pady=5)
    ttk.Button(win, text="Închide", command=win.destroy).grid(row=1, column=3, pady=5)

    refresh()


def open_manage_window(root):
    """Simple panel with admin actions."""

    win = tk.Toplevel(root)
    win.title("Administrare aplicație")

    frm = ttk.Frame(win, padding=10)
    frm.grid(sticky="nsew")
    win.columnconfigure(0, weight=1)
    win.rowconfigure(0, weight=1)

    btns = [
        ("Utilizatori", lambda: open_users_window(win)),
        ("Firme facturare", lambda: open_firme_window(win)),
        ("Locații", root.lift),
        ("Închide", win.destroy),
    ]
    for i, (txt, cmd) in enumerate(btns):
        ttk.Button(frm, text=txt, command=cmd).grid(row=i, column=0, sticky="ew", pady=5)
    frm.columnconfigure(0, weight=1)

