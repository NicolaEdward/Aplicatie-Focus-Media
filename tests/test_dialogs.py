import pandas as pd
import UI.dialogs as dialogs
import db


def test_export_sales_report_year_no_reservations(monkeypatch, tmp_path):
    # Avoid database and UI interactions
    monkeypatch.setattr(db, "update_statusuri_din_rezervari", lambda: None)
    monkeypatch.setattr(dialogs, "choose_report_year", lambda parent=None: 2030)

    def dummy_read_sql_query(sql, params=None, **kwargs):
        if "FROM locatii" in sql:
            return pd.DataFrame({
                "id": [1],
                "city": ["c"],
                "county": ["co"],
                "address": ["a"],
                "type": ["t"],
                "size": ["s"],
                "sqm": [10],
                "illumination": ["i"],
                "ratecard": [100],
                "pret_vanzare": [1000],
                "grup": ["g"],
                "status": ["free"],
                "is_mobile": [0],
                "parent_id": [None],
            })
        return pd.DataFrame(
            columns=[
                "id",
                "grup",
                "city",
                "county",
                "address",
                "type",
                "size",
                "sqm",
                "illumination",
                "ratecard",
                "pret_vanzare",
                "is_mobile",
                "parent_id",
                "client",
                "data_start",
                "data_end",
                "suma",
            ]
        )

    monkeypatch.setattr(db, "read_sql_query", dummy_read_sql_query)

    saved = {}
    monkeypatch.setattr(dialogs.filedialog, "asksaveasfilename", lambda **k: str(tmp_path / "out.xlsx"))
    monkeypatch.setattr(dialogs.messagebox, "showinfo", lambda *a, **k: saved.setdefault("info", a))

    dialogs.export_sales_report()
    assert (tmp_path / "out.xlsx").exists()
    assert saved.get("info")


def test_export_client_backup(monkeypatch, tmp_path):
    row = (
        "Cli",
        "RO1",
        "AddrC",
        "Firma",
        "FCUI",
        "AddrF",
        "Camp",
        "City",
        "Addr",
        "CODE",
        "F1",
        "Billboard",
        "5x3",
        15,
        "2023-05-01",
        "2023-05-31",
        100.0,
        10.0,
        1,
        None,
        None,
        1,
        1,
    )

    class DummyCursor:
        def __init__(self):
            self.call = 0

        def execute(self, sql, params=()):
            self.call += 1
            return self

        def fetchall(self):
            return [row]

        def fetchone(self):
            return (0.0, 0.0)

    class DummyConn:
        def cursor(self):
            return DummyCursor()

    dummy = DummyConn()
    monkeypatch.setattr(dialogs, "conn", dummy)

    monkeypatch.setattr(dialogs.filedialog, "askdirectory", lambda **k: str(tmp_path))
    saved = {}
    monkeypatch.setattr(dialogs.messagebox, "showinfo", lambda *a, **k: saved.setdefault("info", a))

    dialogs.export_client_backup(5, 2023, 1)
    month_dir = tmp_path / "BKP May 2023"
    expected = month_dir / "Firma" / "BKP Firma x Cli - Camp - May.xlsx"
    assert expected.exists()
    assert saved.get("info")


def test_export_all_backups(monkeypatch, tmp_path):
    rows = [
        (
            "Cli",
            "RO1",
            "AddrC",
            "Firma",
            "FCUI",
            "AddrF",
            "Camp1",
            "City",
            "Addr",
            "CODE",
            "F1",
            "Billboard",
            "5x3",
            15,
            "2023-05-01",
            "2023-05-31",
            100.0,
            10.0,
            1,
            None,
            None,
            1,
            1,
        ),
        (
            "Cli",
            "RO1",
            "AddrC",
            "Firma2",
            "FCUI",
            "AddrF",
            "Camp2",
            "City",
            "Addr",
            "CODE",
            "F1",
            "Billboard",
            "5x3",
            15,
            "2023-05-01",
            "2023-05-31",
            100.0,
            10.0,
            1,
            None,
            None,
            2,
            1,
        ),
    ]

    class DummyCursor:
        def __init__(self):
            self.call = 0

        def execute(self, sql, params=()):
            self.call += 1
            return self

        def fetchall(self):
            return rows

        def fetchone(self):
            return (0.0, 0.0)

    class DummyConn:
        def cursor(self):
            return DummyCursor()

    dummy = DummyConn()
    monkeypatch.setattr(dialogs, "conn", dummy)

    monkeypatch.setattr(dialogs.filedialog, "askdirectory", lambda **k: str(tmp_path))
    saved = {}
    monkeypatch.setattr(dialogs.messagebox, "showinfo", lambda *a, **k: saved.setdefault("info", a))

    dialogs.export_all_backups(5, 2023)
    month_dir = tmp_path / "BKP May 2023"
    assert (month_dir / "Firma" / "BKP Firma x Cli - Camp1 - May.xlsx").exists()
    assert (month_dir / "Firma2" / "BKP Firma2 x Cli - Camp2 - May.xlsx").exists()
    assert saved.get("info")


def test_vendor_report_created_in_month(monkeypatch, tmp_path):
    import datetime
    from openpyxl import load_workbook

    month = 6
    year = 2025

    users_df = pd.DataFrame({"username": ["s1"], "comune": [""]})
    rezerv_df = pd.DataFrame(
        {
            "created_by": ["s1", "s1", "s1"],
            "created_on": [
                datetime.datetime(2025, 5, 1),
                datetime.datetime(2025, 6, 5),
                datetime.datetime(2025, 6, 10),
            ],
            "client": ["CL1", "CL2", "CL3"],
            "suma": [1000.0, 2000.0, 1500.0],
            "data_start": [
                datetime.datetime(2025, 5, 1),
                datetime.datetime(2025, 5, 20),
                datetime.datetime(2025, 6, 20),
            ],
            "data_end": [
                datetime.datetime(2025, 5, 31),
                datetime.datetime(2025, 8, 1),
                datetime.datetime(2025, 7, 10),
            ],
            "city": ["C1", "C2", "C3"],
            "county": ["Co", "Co", "Co"],
            "address": ["A1", "A2", "A3"],
        }
    )

    def dummy_read_sql_query(sql, params=None, **kwargs):
        if "FROM users" in sql:
            return users_df
        return rezerv_df

    monkeypatch.setattr(db, "read_sql_query", dummy_read_sql_query)
    monkeypatch.setattr(dialogs.filedialog, "asksaveasfilename", lambda **k: str(tmp_path / "out.xlsx"))
    saved = {}
    monkeypatch.setattr(dialogs.messagebox, "showinfo", lambda *a, **k: saved.setdefault("info", a))

    dialogs.export_vendor_report(month=month, year=year)
    assert saved.get("info")

    wb = load_workbook(tmp_path / "out.xlsx")
    ws = wb["s1"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 3  # header, 1 entry, total
    assert rows[1][0] == 1
    assert rows[2][4] == "Total"

    saved.clear()
    monkeypatch.setattr(dialogs.filedialog, "asksaveasfilename", lambda **k: str(tmp_path / "out2.xlsx"))
    dialogs.export_vendor_report(month=7, year=year)
    assert "Nu există" in saved.get("info")[1]

